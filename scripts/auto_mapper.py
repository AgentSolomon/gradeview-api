#!/usr/bin/env python3
"""
GradeView — Column Auto-Mapper
================================
Inspects xlsx/csv headers and generates a MAPPINGS-compatible profile
for upload_v2.py. Handles both standard (one row per grade) and wide
(one row per section, grade columns) formats.

Usage:
    # Inspect a file and print proposed mapping
    python3 auto_mapper.py --file data.xlsx

    # Inspect and auto-generate a mapping for upload_v2.py
    python3 auto_mapper.py --file data.xlsx --school gvsu

    # Inspect with verbose output (shows all columns + reasoning)
    python3 auto_mapper.py --file data.xlsx --verbose

    # Output JSON mapping (for programmatic use)
    python3 auto_mapper.py --file data.xlsx --json

    # Just print headers and sample data (no mapping attempt)
    python3 auto_mapper.py --file data.xlsx --inspect-only

Called from upload_v2.py:
    from auto_mapper import auto_map
    mapping = auto_map("path/to/file.xlsx")
"""

import argparse
import csv
import json
import os
import re
import sys
from collections import Counter

# ── Known Grade Values ───────────────────────────────────────────────────────
# Column names that match these are almost certainly grade-count columns in
# wide-format data. Kept as a set for O(1) lookup.

KNOWN_GRADES = {
    # Standard letter grades
    "A+", "A", "A-",
    "B+", "B", "B-",
    "C+", "C", "C-",
    "D+", "D", "D-",
    "F", "W", "I", "P", "NP",
    # Common extras
    "AU", "CR", "NC", "X", "NR", "IP", "WF", "WP", "S", "U",
    # Underscore-suffixed (UNLV/NEIU style)
    "A_GRADE", "AMINUS_GRADE", "BPLUS_GRADE", "B_GRADE", "BMINUS_GRADE",
    "CPLUS_GRADE", "C_GRADE", "CMINUS_GRADE", "DPLUS_GRADE", "D_GRADE",
    "DMINUS_GRADE", "F_GRADE", "W_GRADE", "I_GRADE", "PASS_GRADE",
}

# Grade column names that map to letter grades (for underscore-style columns)
GRADE_COL_TO_LETTER = {
    "A_GRADE": "A", "AMINUS_GRADE": "A-", "APLUS_GRADE": "A+",
    "BPLUS_GRADE": "B+", "B_GRADE": "B", "BMINUS_GRADE": "B-",
    "CPLUS_GRADE": "C+", "C_GRADE": "C", "CMINUS_GRADE": "C-",
    "DPLUS_GRADE": "D+", "D_GRADE": "D", "DMINUS_GRADE": "D-",
    "F_GRADE": "F", "W_GRADE": "W", "I_GRADE": "I",
    "PASS_GRADE": "P", "NP_GRADE": "NP",
}

# ── Semantic Field Matchers ──────────────────────────────────────────────────
# Each field has a list of (pattern, weight) tuples. Patterns are matched
# against column names (case-insensitive). Higher weight = more confident.

FIELD_PATTERNS = {
    "semester": [
        (r"^semester$", 100),
        (r"^term$", 95),
        (r"^acad(emic)?[\s_]?term$", 90),
        (r"^acad(emic)?[\s_]?period$", 80),
        (r"^session$", 60),
        (r"semester", 50),
        (r"term", 40),
    ],
    "year": [
        (r"^year$", 100),
        (r"^acad(emic)?[\s_]?year$", 95),
        (r"^calendar[\s_]?year$", 90),
        (r"^fiscal[\s_]?year$", 70),
        (r"year", 50),
    ],
    "dept": [
        (r"^dept$", 100),
        (r"^department$", 100),
        (r"^subject$", 90),
        (r"^course[\s_]?prefix$", 90),
        (r"^course[\s_]?subject$", 85),
        (r"^discipline$", 80),
        (r"^program$", 60),
        (r"subject", 50),
        (r"dept", 50),
        (r"prefix", 40),
    ],
    "course_number": [
        (r"^course[\s_]?(number|no|num|nbr)$", 100),
        (r"^course$", 85),
        (r"^catalog[\s_]?(number|no|num|nbr)$", 90),
        (r"^class[\s_]?(number|no|num|nbr)$", 85),
        (r"^course[\s_]?code$", 80),
        (r"^section$", 70),  # often compound: "AAA 200 01"
        (r"course", 40),
    ],
    "instructor": [
        (r"^instructor$", 100),
        (r"^instructor[\s_]?name$", 100),
        (r"^professor$", 95),
        (r"^faculty$", 90),
        (r"^teacher$", 85),
        (r"^primary[\s_]?instructor", 90),
        (r"instructor", 50),
        (r"professor", 50),
        (r"faculty", 40),
    ],
    "instructor_first": [
        (r"^(primary[\s_]?)?instructor[\s_]?first", 100),
        (r"^faculty[\s_]?first", 90),
        (r"^first[\s_]?name$", 50),
    ],
    "instructor_last": [
        (r"^(primary[\s_]?)?instructor[\s_]?last", 100),
        (r"^faculty[\s_]?last", 90),
        (r"^last[\s_]?name$", 50),
    ],
    "grade": [
        (r"^grade$", 100),
        (r"^letter[\s_]?grade$", 100),
        (r"^final[\s_]?grade$", 95),
        (r"^grade[\s_]?letter$", 90),
        (r"grade", 50),
    ],
    "count": [
        (r"^count$", 100),
        (r"^count[\s_]?of[\s_]?letter[\s_]?grade$", 100),
        (r"^n$", 80),
        (r"^enrollment$", 75),
        (r"^frequency$", 70),
        (r"^num[\s_]?students$", 70),
        (r"count", 50),
        (r"enrollment", 40),
    ],
}

# Columns to always skip (metadata, not grade data)
SKIP_PATTERNS = [
    r"^total$", r"^gpa$", r"^avg", r"^mean", r"^median",
    r"^title$", r"^course[\s_]?title$", r"^class[\s_]?title$",
    r"^section[\s_]?(number|no|num)$",
    r"^crn$", r"^id$", r"^row", r"^index",
    r"^campus$", r"^college$", r"^level$",
]


# ── File Reading (minimal, for header inspection) ────────────────────────────

def read_headers_and_sample(filepath, sheet_name=None, sample_rows=10):
    """Read column headers and a few sample rows from a CSV or XLSX file."""
    if filepath.endswith(('.xlsx', '.xls')):
        return _read_xlsx_sample(filepath, sheet_name, sample_rows)
    else:
        return _read_csv_sample(filepath, sample_rows)

def _read_xlsx_sample(filepath, sheet_name, sample_rows):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheets_info = wb.sheetnames
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    active_sheet = ws.title

    headers = None
    rows = []
    for excel_row in ws.iter_rows(max_row=sample_rows + 1, values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else '' for c in excel_row]
            continue
        row = dict(zip(headers, [str(c).strip() if c is not None else '' for c in excel_row]))
        rows.append(row)
    wb.close()

    # Get total row count (rough estimate from read_only)
    wb2 = openpyxl.load_workbook(filepath, read_only=True)
    ws2 = wb2[active_sheet]
    total = sum(1 for _ in ws2.iter_rows(values_only=True)) - 1  # minus header
    wb2.close()

    return {
        "headers": headers,
        "sample": rows,
        "total_rows": total,
        "sheets": sheets_info,
        "active_sheet": active_sheet,
        "filepath": filepath,
    }

def _read_csv_sample(filepath, sample_rows):
    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = []
        total = 0
        for row in reader:
            total += 1
            if len(rows) < sample_rows:
                rows.append(row)
    return {
        "headers": list(headers),
        "sample": rows,
        "total_rows": total,
        "sheets": None,
        "active_sheet": None,
        "filepath": filepath,
    }


# ── Format Detection ────────────────────────────────────────────────────────

def detect_format(headers, sample_rows):
    """
    Detect whether the data is 'standard' (one row per grade entry with a
    grade column and a count column) or 'wide' (one row per section with
    individual grade columns like A, B+, etc.).

    Returns: ("standard", grade_cols=[]) or ("wide", grade_cols=[...])
    """
    grade_cols = []
    for h in headers:
        h_upper = h.strip().upper()
        h_clean = h.strip()
        if h_clean in KNOWN_GRADES or h_upper in KNOWN_GRADES:
            grade_cols.append(h_clean)

    # If we find 3+ grade-like column names, it's almost certainly wide format
    if len(grade_cols) >= 3:
        # Verify by checking if the values are numeric
        numeric_count = 0
        for row in sample_rows:
            for gc in grade_cols[:5]:
                val = row.get(gc, '')
                try:
                    int(float(val)) if val else None
                    numeric_count += 1
                except (ValueError, TypeError):
                    pass
        # If most grade-col values are numeric, confirmed wide format
        if numeric_count >= len(sample_rows):
            return "wide", grade_cols

    return "standard", []


# ── Column Matching ──────────────────────────────────────────────────────────

def match_field(header, field_name):
    """Score how well a header matches a field. Returns 0 if no match."""
    h_lower = header.lower().strip()
    for pattern, weight in FIELD_PATTERNS.get(field_name, []):
        if re.search(pattern, h_lower):
            return weight
    return 0

def is_skip_column(header):
    """Check if a column should be skipped (metadata, not useful for mapping)."""
    h_lower = header.lower().strip()
    for pattern in SKIP_PATTERNS:
        if re.search(pattern, h_lower):
            return True
    return False

def detect_compound_section(headers, sample_rows):
    """
    Detect if a 'Section' column contains compound data like 'AAA 200 01'
    that encodes dept + course_number + section_number.

    Returns: (column_name, parse_function) or (None, None)
    """
    section_col = None
    for h in headers:
        if re.search(r"^section$", h.lower().strip()):
            section_col = h
            break

    if not section_col:
        return None, None

    # Check sample values for the pattern: ALPHA DIGITS [DIGITS]
    compound_count = 0
    for row in sample_rows:
        val = str(row.get(section_col, '')).strip()
        # Match patterns like "AAA 200 01", "CS 101", "MATH 2413 001"
        if re.match(r'^[A-Z]{2,6}\s+\d{3,4}(\s+\d{1,3})?$', val):
            compound_count += 1

    if compound_count >= len(sample_rows) * 0.7:  # 70% threshold
        return section_col, _parse_compound_section

    return None, None

def _parse_compound_section(value):
    """Parse 'AAA 200 01' into (dept='AAA', course='200')."""
    parts = str(value).strip().split()
    if len(parts) >= 2:
        return parts[0], parts[1]  # dept, course_number
    return value, ''


# ── Auto-Mapping Engine ─────────────────────────────────────────────────────

def auto_map(filepath, sheet_name=None, school_id=None, verbose=False):
    """
    Inspect a file and generate a MAPPINGS-compatible profile.

    Returns a dict with:
        - "mapping": the generated mapping dict (MAPPINGS-compatible)
        - "format": "standard" or "wide"
        - "confidence": float 0-1 (how confident we are in the mapping)
        - "warnings": list of issues found
        - "unmapped_headers": headers we couldn't classify
        - "info": file metadata (rows, sheets, etc.)
    """
    info = read_headers_and_sample(filepath, sheet_name)
    headers = info["headers"]
    sample = info["sample"]
    warnings = []

    if not headers:
        return {
            "mapping": None, "format": None, "confidence": 0.0,
            "warnings": ["No headers found in file"],
            "unmapped_headers": [], "info": info,
        }

    # Step 1: Detect format
    fmt, grade_cols = detect_format(headers, sample)

    # Step 2: Check for compound section column
    compound_col, compound_parser = detect_compound_section(headers, sample)

    # Step 3: Match each non-grade, non-skip header to a field
    field_scores = {}  # field_name -> [(header, score)]
    for h in headers:
        if not h.strip():
            continue
        if h in grade_cols:
            continue
        if is_skip_column(h):
            continue

        for field_name in FIELD_PATTERNS:
            score = match_field(h, field_name)
            if score > 0:
                if field_name not in field_scores:
                    field_scores[field_name] = []
                field_scores[field_name].append((h, score))

    # Step 4: Pick best match for each field (highest score wins)
    assigned = {}  # field_name -> header
    used_headers = set()

    # Sort fields by max available score (most confident first)
    field_order = sorted(
        field_scores.keys(),
        key=lambda f: max(s for _, s in field_scores[f]),
        reverse=True,
    )

    for field_name in field_order:
        candidates = sorted(field_scores[field_name], key=lambda x: -x[1])
        for header, score in candidates:
            if header not in used_headers:
                assigned[field_name] = header
                used_headers.add(header)
                break

    # Step 5: Handle compound section override
    if compound_col:
        # If we matched section to course_number, override with compound logic
        if assigned.get("course_number") == compound_col:
            assigned["_compound_section"] = compound_col
            # Don't overwrite course_number — we'll parse it from section
            if verbose:
                warnings.append(
                    f"Compound 'Section' column detected: '{compound_col}' "
                    f"→ will split into dept + course_number"
                )

    # Step 6: Build the mapping dict
    mapping = {}

    if fmt == "wide":
        mapping["format"] = "wide"

        if info.get("active_sheet"):
            mapping["sheet"] = info["active_sheet"]

        # Year: check if semester contains year or if there's a separate year col
        if "year" in assigned:
            mapping["year"] = assigned["year"]
        elif "semester" in assigned:
            # Check if semester values contain years (e.g., "Fall 2020")
            for row in sample:
                sem_val = row.get(assigned["semester"], '')
                if re.search(r'\b(19|20)\d{2}\b', str(sem_val)):
                    mapping["_year_in_semester"] = True
                    break
            if "_year_in_semester" not in mapping:
                warnings.append("No year column found and semester doesn't contain year")

        # Semester
        if "semester" in assigned:
            mapping["semester"] = assigned["semester"]
        else:
            warnings.append("No semester/term column found")

        # Dept
        if compound_col and compound_col in used_headers:
            mapping["_compound_section"] = compound_col
            mapping["dept"] = None  # will be parsed from section
            mapping["course_number"] = None
        elif "dept" in assigned:
            mapping["dept"] = assigned["dept"]
        else:
            warnings.append("No department/subject column found")

        # Course number
        if not compound_col or compound_col not in used_headers:
            if "course_number" in assigned:
                mapping["course_number"] = assigned["course_number"]
            else:
                warnings.append("No course number column found")

        # Instructor
        if "instructor" in assigned:
            mapping["instructor_first"] = assigned["instructor"]
            mapping["instructor_last"] = None
        elif "instructor_first" in assigned:
            mapping["instructor_first"] = assigned["instructor_first"]
            mapping["instructor_last"] = assigned.get("instructor_last")
        else:
            warnings.append("No instructor column found — will default to 'N/A'")

        # Grade columns
        grade_col_map = {}
        for gc in grade_cols:
            # Map column name to letter grade
            gc_upper = gc.upper()
            if gc_upper in GRADE_COL_TO_LETTER:
                grade_col_map[gc] = GRADE_COL_TO_LETTER[gc_upper]
            elif gc in ("A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-",
                        "D+", "D", "D-", "F", "W", "I", "P", "NP",
                        "AU", "CR", "NC", "X", "NR", "IP", "WF", "WP", "S", "U"):
                grade_col_map[gc] = gc
            else:
                grade_col_map[gc] = gc  # pass through
        mapping["grade_cols"] = grade_col_map

    else:
        # Standard format
        for field in ("semester", "year", "dept", "course_number", "grade", "count", "instructor"):
            if field in assigned:
                mapping[field] = assigned[field]
            elif field == "year":
                # Year is optional — can be embedded in semester column instead
                pass
            elif field == "instructor":
                mapping[field] = None
                warnings.append("No instructor column found — will default to 'N/A'")
            else:
                warnings.append(f"No match found for required field: {field}")

        if "instructor_first" in assigned and "instructor" not in assigned:
            mapping["instructor"] = assigned["instructor_first"]

    # Step 7: Identify unmapped headers
    unmapped = [h for h in headers if h.strip() and h not in used_headers and h not in grade_cols]
    # Filter out known skip columns
    unmapped = [h for h in unmapped if not is_skip_column(h)]

    # Step 8: Calculate confidence
    if fmt == "wide":
        required = ["semester", "grade_cols"]
        need_dept = "dept" in mapping or "_compound_section" in mapping
        have_grades = len(mapping.get("grade_cols", {})) >= 3
        conf_parts = [
            0.3 if "semester" in mapping else 0.0,
            0.3 if need_dept else 0.0,
            0.3 if have_grades else 0.0,
            0.1 if ("instructor_first" in mapping or "instructor" in assigned) else 0.0,
        ]
    else:
        conf_parts = [
            0.2 if "semester" in mapping else 0.0,
            0.15 if "dept" in mapping else 0.0,
            0.15 if "course_number" in mapping else 0.0,
            0.2 if "grade" in mapping else 0.0,
            0.2 if "count" in mapping else 0.0,
            0.1 if mapping.get("instructor") else 0.0,
        ]
    confidence = sum(conf_parts)

    return {
        "mapping": mapping,
        "format": fmt,
        "confidence": confidence,
        "warnings": warnings,
        "unmapped_headers": unmapped,
        "info": {
            "total_rows": info["total_rows"],
            "sheets": info["sheets"],
            "active_sheet": info["active_sheet"],
            "headers": info["headers"],
            "filepath": info["filepath"],
        },
        "sample": sample[:3],
        "compound_section": compound_col,
    }


def build_upload_mapping(result):
    """
    Convert auto_map result into a mapping dict that upload_v2.py can use
    directly in its MAPPINGS dictionary. Handles compound sections by
    adding parsing metadata.
    """
    m = result["mapping"]
    if not m:
        return None

    # Clean copy without internal keys
    clean = {k: v for k, v in m.items() if not k.startswith("_")}

    # If compound section detected, add parsing instructions
    if result.get("compound_section"):
        clean["_parse_section"] = {
            "column": result["compound_section"],
            "dept_index": 0,
            "course_index": 1,
        }

    return clean


# ── Pretty Printer ───────────────────────────────────────────────────────────

def print_report(result, school_id=None, verbose=False):
    """Print a human-readable mapping report."""
    info = result["info"]
    mapping = result["mapping"]
    fmt = result["format"]
    conf = result["confidence"]

    print(f"\n{'='*60}")
    print(f"  AUTO-MAPPER REPORT")
    print(f"{'='*60}")
    print(f"  File:       {os.path.basename(info['filepath'])}")
    print(f"  Rows:       {info['total_rows']:,}")
    if info['sheets']:
        print(f"  Sheets:     {', '.join(info['sheets'])}")
        print(f"  Active:     {info['active_sheet']}")
    print(f"  Format:     {fmt.upper()}")
    print(f"  Confidence: {conf:.0%}")
    if school_id:
        print(f"  School ID:  {school_id}")

    print(f"\n  Headers ({len(info['headers'])}):")
    for i, h in enumerate(info['headers']):
        print(f"    [{i:2d}] {h}")

    print(f"\n  Proposed Mapping:")
    if fmt == "wide":
        for key in ("sheet", "year", "semester", "dept", "course_number",
                     "instructor_first", "instructor_last"):
            if key in mapping:
                print(f"    {key:20s} → {mapping[key]}")
        if result.get("compound_section"):
            print(f"    {'(compound section)':20s} → '{result['compound_section']}' "
                  f"→ split into dept + course_number")
        if "_year_in_semester" in mapping:
            print(f"    {'(year)':20s} → parsed from semester column")
        if "grade_cols" in mapping:
            gc = mapping["grade_cols"]
            print(f"    grade_cols ({len(gc)}):")
            for col, letter in sorted(gc.items(), key=lambda x: x[1]):
                print(f"      {col:20s} → {letter}")
    else:
        for key in ("semester", "dept", "course_number", "grade", "count", "instructor"):
            if key in mapping:
                print(f"    {key:20s} → {mapping[key]}")

    if result["warnings"]:
        print(f"\n  Warnings:")
        for w in result["warnings"]:
            print(f"    ⚠️  {w}")

    if result["unmapped_headers"]:
        print(f"\n  Unmapped columns (ignored):")
        for h in result["unmapped_headers"]:
            print(f"    ·  {h}")

    if result.get("sample"):
        print(f"\n  Sample data (first {len(result['sample'])} rows):")
        for i, row in enumerate(result["sample"]):
            vals = [f"{k}={v}" for k, v in list(row.items())[:6]]
            print(f"    [{i}] {', '.join(vals)}")

    print(f"\n{'='*60}")

    if conf >= 0.8:
        print(f"  ✅ HIGH CONFIDENCE — ready for staging")
    elif conf >= 0.5:
        print(f"  ⚠️  MEDIUM CONFIDENCE — review mapping before staging")
    else:
        print(f"  ❌ LOW CONFIDENCE — manual mapping likely needed")

    # Print copy-pasteable MAPPINGS entry
    if school_id and mapping:
        upload_mapping = build_upload_mapping(result)
        print(f"\n  Copy-paste for upload_v2.py MAPPINGS dict:")
        print(f'    "{school_id}": {json.dumps(upload_mapping, indent=8)},')

    print()


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='GradeView Column Auto-Mapper — inspect files and generate upload mappings'
    )
    parser.add_argument('--file', required=True, help='Path to CSV/XLSX file')
    parser.add_argument('--school', help='School ID (for mapping output)')
    parser.add_argument('--sheet', help='Specific sheet name to inspect')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    parser.add_argument('--json', action='store_true', help='Output JSON mapping')
    parser.add_argument('--inspect-only', action='store_true', help='Just show headers and sample')

    args = parser.parse_args()

    filepath = os.path.expanduser(args.file)
    if not os.path.exists(filepath):
        print(f"❌ File not found: {filepath}")
        sys.exit(1)

    if args.inspect_only:
        info = read_headers_and_sample(filepath, args.sheet)
        print(f"\nFile: {os.path.basename(filepath)}")
        print(f"Rows: {info['total_rows']:,}")
        if info['sheets']:
            print(f"Sheets: {', '.join(info['sheets'])}")
        print(f"\nHeaders ({len(info['headers'])}):")
        for i, h in enumerate(info['headers']):
            print(f"  [{i:2d}] {h}")
        print(f"\nSample ({len(info['sample'])} rows):")
        for i, row in enumerate(info['sample'][:5]):
            print(f"  [{i}] {dict(list(row.items())[:8])}")
        return

    result = auto_map(filepath, sheet_name=args.sheet, school_id=args.school,
                      verbose=args.verbose)

    if args.json:
        output = {
            "school_id": args.school,
            "format": result["format"],
            "confidence": result["confidence"],
            "mapping": build_upload_mapping(result),
            "warnings": result["warnings"],
            "unmapped_headers": result["unmapped_headers"],
            "total_rows": result["info"]["total_rows"],
        }
        print(json.dumps(output, indent=2))
    else:
        print_report(result, school_id=args.school, verbose=args.verbose)


if __name__ == '__main__':
    main()
