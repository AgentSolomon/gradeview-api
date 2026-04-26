#!/usr/bin/env python3
"""
University of Iowa grade distribution extractor.
Input:  ~/Documents/GradeView/raw_data/uiowa/uiowa_extracted/Grade distribution reports/*.xlsx
Output: ~/Documents/GradeView/raw_data/uiowa/uiowa_grades_extracted.csv

Format: Each row is a course section with grade RANGE columns:
  +A-  = A grades (A+, A, A-)
  +B-  = B grades
  +C-  = C grades
  +D-_F = D and F grades combined
  WITHDRAWN = W

We expand each range bucket into a single row per grade bucket.
SUBJECT_COURSE_SECTION format: DEPT:COURSE:SECTION (e.g. ACCT:2100:0001)
INSTRUCTOR_NAME_AND_ROLE: "Firstname Lastname - Role"
Sheet name encodes semester: "Summer2023final" → year=2023, semester=SUMMER
"""

import csv
import glob
import re
import sys
from pathlib import Path
import openpyxl

RAW_DIR  = Path.home() / "Documents/GradeView/raw_data/uiowa/uiowa_extracted/Grade distribution reports"
OUTPUT   = Path.home() / "Documents/GradeView/raw_data/uiowa/uiowa_grades_extracted.csv"

GRADE_COL_MAP = {
    "+A-":    "A",
    "+B-":    "B",
    "+C-":    "C",
    "+D-_F":  "D/F",
    "WITHDRAWN": "W",
}

SEMESTER_MAP = {
    "summer": "SUMMER",
    "fall":   "FALL",
    "winter": "WINTER",
    "spring": "SPRING",
}

def parse_sheet_name(name):
    """'Summer2023final' → (2023, SUMMER)"""
    name_lower = name.lower()
    for key, sem in SEMESTER_MAP.items():
        if key in name_lower:
            m = re.search(r'(\d{4})', name)
            year = m.group(1) if m else "0000"
            return year, sem
    return "0000", "UNKNOWN"

def parse_instructor(raw):
    """'Christian Hutzler - Primary Instructor' → 'Hutzler, Christian'"""
    if not raw:
        return "N/A"
    raw = str(raw).strip()
    # Strip role suffix
    parts = raw.split(" - ")
    name = parts[0].strip()
    # Convert "Firstname Lastname" → "Lastname, Firstname"
    name_parts = name.split()
    if len(name_parts) >= 2:
        return f"{name_parts[-1]}, {' '.join(name_parts[:-1])}"
    return name

def parse_subject_course(scs):
    """'ACCT:2100:0001' → (dept='ACCT', course='2100')"""
    parts = str(scs).split(":")
    dept   = parts[0].strip() if len(parts) > 0 else ""
    course = parts[1].strip() if len(parts) > 1 else ""
    return dept, course

def extract_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out = []

    for sheet_name in wb.sheetnames:
        year, semester = parse_sheet_name(sheet_name)
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        headers = [str(h).strip() if h else "" for h in all_rows[0]]

        try:
            i_scs  = headers.index("SUBJECT_COURSE_SECTION")
            i_inst = headers.index("INSTRUCTOR_NAME_AND_ROLE")
        except ValueError as e:
            print(f"  ⚠️  Skipping sheet '{sheet_name}': {e}")
            continue

        # Find grade columns
        grade_cols = {}
        for col_name, grade_label in GRADE_COL_MAP.items():
            if col_name in headers:
                grade_cols[headers.index(col_name)] = grade_label

        if not grade_cols:
            print(f"  ⚠️  No grade columns found in '{sheet_name}'")
            continue

        for row in all_rows[1:]:
            scs = row[i_scs]
            if not scs:
                continue
            dept, course = parse_subject_course(scs)
            instructor = parse_instructor(row[i_inst])

            for col_idx, grade_label in grade_cols.items():
                count = row[col_idx]
                try:
                    count = int(count)
                except (TypeError, ValueError):
                    count = 0
                if count <= 0:
                    continue
                rows_out.append({
                    "school_id":     "uiowa",
                    "year":          year,
                    "semester":      semester,
                    "dept":          dept,
                    "course_number": course,
                    "instructor":    instructor,
                    "grade":         grade_label,
                    "count":         count,
                })

    wb.close()
    return rows_out

def main():
    files = sorted(glob.glob(str(RAW_DIR / "*.xlsx")))
    if not files:
        print("No xlsx files found in", RAW_DIR)
        sys.exit(1)

    all_rows = []
    for f in files:
        print(f"  Processing {Path(f).name}...")
        rows = extract_file(f)
        print(f"    → {len(rows):,} rows")
        all_rows.extend(rows)

    print(f"\nTotal rows: {len(all_rows):,}")

    fieldnames = ["school_id","year","semester","dept","course_number","instructor","grade","count"]
    with open(OUTPUT, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"Saved → {OUTPUT}")

if __name__ == "__main__":
    main()
