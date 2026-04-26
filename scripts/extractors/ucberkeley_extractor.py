#!/usr/bin/env python3
"""UC Berkeley extractor — wide format, grades as columns, header on row 11, semester in cell A1 area."""
import csv, glob, re
from pathlib import Path
from collections import defaultdict
import openpyxl

RAW_DIR = Path.home() / "Documents/GradeView/raw_data/ucberkeley"
OUTPUT  = RAW_DIR / "ucberkeley_grades_extracted.csv"

GRADE_COLS = ['A+','A','A-','B+','B','B-','C+','C','C-','D+','D','D-','F','Pass','Not Pass']

SEMESTER_MAP = {
    'spring': 'SPRING', 'fall': 'FALL', 'summer': 'SUMMER', 'winter': 'WINTER'
}

def parse_sheet_semester(ws, sheet_name):
    """Extract year and semester from sheet name or cell A1."""
    name = sheet_name.lower()
    for key, sem in SEMESTER_MAP.items():
        if key in name:
            m = re.search(r'(\d{4})', sheet_name)
            year = m.group(1) if m else '0000'
            return year, sem
    # Try reading from cell A1 area
    rows = list(ws.iter_rows(max_row=5, values_only=True))
    for row in rows:
        if row[0]:
            txt = str(row[0])
            m = re.search(r'(Spring|Fall|Summer|Winter)\s+(\d{4})', txt, re.I)
            if m:
                return m.group(2), m.group(1).upper()
    return '0000', 'UNKNOWN'

def extract_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ('grade scale', 'notes', 'readme'):
            continue
        ws = wb[sheet_name]
        year, semester = parse_sheet_semester(ws, sheet_name)
        all_rows = list(ws.iter_rows(values_only=True))
        # Find header row (contains 'Course Number' or 'A+')
        header_idx = None
        for i, row in enumerate(all_rows):
            if row and any(str(c) in ('A+', 'Course Number', 'Instr Name Concat') for c in row if c):
                header_idx = i
                break
        if header_idx is None:
            continue
        headers = [str(h).strip() if h else '' for h in all_rows[header_idx]]
        # Find column indices
        try:
            i_dept   = next(i for i,h in enumerate(headers) if 'Subject Cd' in h or 'Subject Short' in h or h == 'Course Subject Cd')
            i_course = next(i for i,h in enumerate(headers) if h == 'Course Number')
            i_instr  = next(i for i,h in enumerate(headers) if 'Instr' in h)
        except StopIteration:
            continue
        # Semester col (optional override)
        i_sem = next((i for i,h in enumerate(headers) if 'Semester' in h or 'Year Name' in h), None)
        grade_idx = {h: i for i,h in enumerate(headers) if h in GRADE_COLS}
        if not grade_idx:
            continue
        for row in all_rows[header_idx+1:]:
            if not row or not row[i_dept]:
                continue
            dept   = str(row[i_dept]).strip()
            course = str(row[i_course]).strip() if row[i_course] else ''
            instr  = str(row[i_instr]).strip() if row[i_instr] else 'N/A'
            # Parse semester from data if available
            row_year, row_sem = year, semester
            if i_sem is not None and row[i_sem]:
                txt = str(row[i_sem])
                m = re.search(r'(Spring|Fall|Summer|Winter)\s+(\d{4})', txt, re.I)
                if m:
                    row_sem  = m.group(1).upper()
                    row_year = m.group(2)
            for grade, col_i in grade_idx.items():
                count = row[col_i]
                try:
                    count = int(count)
                except (TypeError, ValueError):
                    count = 0
                if count <= 0:
                    continue
                rows_out.append({
                    'school_id': 'ucberkeley', 'year': row_year, 'semester': row_sem,
                    'dept': dept, 'course_number': course, 'instructor': instr,
                    'grade': grade, 'count': count
                })
    wb.close()
    return rows_out

def main():
    files = sorted(glob.glob(str(RAW_DIR / "*.xlsx")))
    all_rows = []
    for f in files:
        print(f"  {Path(f).name}...")
        r = extract_file(f)
        print(f"    → {len(r):,} rows")
        all_rows.extend(r)
    print(f"\nTotal: {len(all_rows):,}")
    fieldnames = ['school_id','year','semester','dept','course_number','instructor','grade','count']
    with open(OUTPUT, 'w', newline='', encoding='utf-8') as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)
    print(f"Saved → {OUTPUT}")

if __name__ == '__main__':
    main()
