"""
University of Minnesota grade distribution extractor.
Handles 7 files: TC 2010-2022 (long format), TC 2023 x2, All Campuses 2024-2025 x4.
Long format: one row per (section × grade) → group by (term, subject, catalog_nbr, instructor) and sum.
"""
import os, re, csv
from pathlib import Path
from collections import defaultdict

SCHOOL_ID = 'umn'
OUTPUT = Path.home() / 'Documents/GradeView/raw_data/umn/umn_extracted.csv'

FILES = [
    Path.home() / 'Downloads/TC - grade distribution_Fall2010 to Summer2022.xlsx',
    Path.home() / 'Downloads/TC - grade distribution_Fall2023.xlsx',
    Path.home() / 'Downloads/TC - grade distribution_Spring2023.xlsx',
    Path.home() / 'Downloads/All Campuses - grade distribution_Fall2024.xlsx',
    Path.home() / 'Downloads/All Campuses - grade_distribution_Spring2024.xlsx',
    Path.home() / 'Downloads/All Campuses - grade distribution_Spring2025.xlsx',
    Path.home() / 'Downloads/All Campuses - grade distribution_Fall2025.xlsx',
]

GRADE_MAP = {
    'A+': 'A', 'A': 'A', 'A-': 'A',
    'B+': 'B', 'B': 'B', 'B-': 'B',
    'C+': 'C', 'C': 'C', 'C-': 'C',
    'D+': 'D', 'D': 'D', 'D-': 'D',
    'F': 'F', 'N': 'F',
    'W': 'W', 'WF': 'W',
    'S': 'S', 'P': 'P',
}
KEEP_GRADES = {'A', 'B', 'C', 'D', 'F', 'W'}

SEM_ALIASES = {
    'Fall': 'Fall', 'Fal': 'Fall',
    'Spring': 'Spring', 'Spr': 'Spring', 'Spg': 'Spring',
    'Summer': 'Summer', 'Sum': 'Summer',
    'Winter': 'Winter', 'Win': 'Winter',
}

def parse_semester(term_descr):
    """'Fall 2023' or 'Spr 2023' → ('Spring', 2023)"""
    parts = str(term_descr).strip().rsplit(' ', 1)
    if len(parts) == 2:
        sem_raw, yr_raw = parts
        sem = SEM_ALIASES.get(sem_raw.strip())
        try:
            return sem, int(yr_raw)
        except:
            pass
    return None, None

def clean_instructor(name):
    """'Ho,Karen' or 'Ho, Karen' → 'Karen Ho'"""
    if not name or str(name).strip() in ('', 'nan', 'None'):
        return 'Unknown'
    name = str(name).strip()
    if ',' in name:
        parts = name.split(',', 1)
        last = parts[0].strip()
        first = parts[1].strip()
        return f"{first} {last}"
    return name

import openpyxl

def process_file(path):
    """Process one UMN file. Returns dict: key=(semester,year,subject,catalog,instructor) → grade→count"""
    print(f"  Loading {path.name}...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Grade_Distribution'] if 'Grade_Distribution' in wb.sheetnames else wb.active
    
    headers = None
    records = defaultdict(lambda: defaultdict(int))
    row_count = 0

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else '' for h in row]
            # Find column indices
            try:
                idx_term = headers.index('TERM_DESCR')
                idx_subj = headers.index('SUBJECT')
                idx_cat  = headers.index('CATALOG_NBR')
                idx_grade = headers.index('CRSE_GRADE_OFF')
                idx_count = headers.index('GRADE_HDCNT')
                idx_name = headers.index('NAME')
            except ValueError as e:
                print(f"    ⚠️ Missing column: {e} — skipping file")
                return {}
            continue

        term_raw = row[idx_term]
        subject  = str(row[idx_subj]).strip() if row[idx_subj] else ''
        catalog  = str(row[idx_cat]).strip() if row[idx_cat] else ''
        grade_raw = str(row[idx_grade]).strip() if row[idx_grade] else ''
        count    = row[idx_count]
        name_raw = row[idx_name]

        if not term_raw or not subject or not catalog:
            continue

        sem, yr = parse_semester(term_raw)
        if not sem or not yr:
            continue

        # Only Fall and Spring
        if sem not in ('Fall', 'Spring'):
            continue

        grade = GRADE_MAP.get(grade_raw)
        if grade not in KEEP_GRADES:
            continue

        try:
            count = int(count) if count else 0
        except:
            count = 0

        instructor = clean_instructor(name_raw)
        key = (sem, yr, subject, catalog, instructor)
        records[key][grade] += count
        row_count += 1

    print(f"    → {row_count:,} source rows, {len(records):,} grouped records")
    return records

def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    
    all_records = defaultdict(lambda: defaultdict(int))
    
    for f in FILES:
        if not f.exists():
            print(f"  ⚠️ File not found: {f.name}")
            continue
        recs = process_file(f)
        for key, grades in recs.items():
            for grade, count in grades.items():
                all_records[key][grade] += count

    print(f"\nTotal unique records: {len(all_records):,}")

    out_rows = 0
    with open(OUTPUT, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['school_id', 'semester', 'year', 'department', 'course_number', 'instructor', 'grade', 'count'])
        for (sem, yr, subj, cat, instr), grades in sorted(all_records.items()):
            for grade in KEEP_GRADES:
                count = grades.get(grade, 0)
                if count > 0:
                    writer.writerow([SCHOOL_ID, sem, yr, subj, cat, instr, grade, count])
                    out_rows += 1

    print(f"Output rows: {out_rows:,}")
    print(f"Saved to: {OUTPUT}")

if __name__ == '__main__':
    main()
