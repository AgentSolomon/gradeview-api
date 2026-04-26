#!/usr/bin/env python3
"""
UC Davis grade distribution extractor.
Input:  ~/Documents/GradeView/raw_data/ucdavis/*.xlsx
Output: ~/Documents/GradeView/raw_data/ucdavis/ucdavis_grades_extracted.csv

Column mapping:
  SUBJ        → dept
  CRSE        → course_number
  TERM        → year (first 4 digits) + semester (last 2: 01=Winter,03=Spring,07=Summer,10=Fall)
  GRADE       → grade
  CNTOFGRADE  → count
  ILNAME+IFNAME → instructor ("Lastname, Firstname")

Grades to SKIP (non-letter, admin codes):
  W04, WC, WPC, NP, NS, NG, I, IP, S, U — keep A/B/C/D/F variants + P/W if present
"""

import csv
import glob
import os
import sys
from pathlib import Path
import openpyxl

RAW_DIR = Path.home() / "Documents/GradeView/raw_data/ucdavis"
OUTPUT  = RAW_DIR / "ucdavis_grades_extracted.csv"

TERM_MAP = {
    "01": "WINTER",
    "03": "SPRING",
    "07": "SUMMER",
    "10": "FALL",
}

# Keep only these grade prefixes
VALID_GRADE_PREFIXES = {"A", "B", "C", "D", "F", "P", "W"}

def is_valid_grade(g):
    if not g:
        return False
    g = str(g).strip().upper()
    return g and g[0] in VALID_GRADE_PREFIXES and len(g) <= 2

def parse_term(term):
    """202010 → (2020, FALL)"""
    t = str(term).strip()
    if len(t) == 6:
        year = t[:4]
        month = t[4:]
        semester = TERM_MAP.get(month, f"TERM{month}")
        return year, semester
    return "0000", "UNKNOWN"

def extract_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        try:
            i_subj   = headers.index("SUBJ")
            i_crse   = headers.index("CRSE")
            i_term   = headers.index("TERM")
            i_grade  = headers.index("GRADE")
            i_cnt    = headers.index("CNTOFGRADE")
            i_lname  = headers.index("ILNAME")
            i_fname  = headers.index("IFNAME")
        except ValueError as e:
            print(f"  ⚠️  Skipping sheet '{sheet_name}': missing column {e}")
            continue

        for row in rows[1:]:
            grade = str(row[i_grade]).strip() if row[i_grade] else ""
            if not is_valid_grade(grade):
                continue
            year, semester = parse_term(row[i_term])
            lname = str(row[i_lname]).strip() if row[i_lname] else ""
            fname = str(row[i_fname]).strip() if row[i_fname] else ""
            instructor = f"{lname}, {fname}".strip(", ") if lname or fname else "N/A"
            count = row[i_cnt]
            try:
                count = int(count)
            except (TypeError, ValueError):
                count = 0
            if count <= 0:
                continue
            rows_out.append({
                "school_id":  "ucdavis",
                "year":       year,
                "semester":   semester,
                "dept":       str(row[i_subj]).strip() if row[i_subj] else "",
                "course_number": str(row[i_crse]).strip() if row[i_crse] else "",
                "instructor": instructor,
                "grade":      grade.upper(),
                "count":      count,
            })
    wb.close()
    return rows_out

def main():
    files = sorted(glob.glob(str(RAW_DIR / "*.xlsx")))
    if not files:
        print("No xlsx files found in", RAW_DIR)
        sys.exit(1)

    all_rows = []
    for f in files:
        print(f"  Processing {Path(f).name}...")
        rows = extract_file(f)
        print(f"    → {len(rows):,} valid rows")
        all_rows.extend(rows)

    print(f"\nTotal rows: {len(all_rows):,}")

    fieldnames = ["school_id","year","semester","dept","course_number","instructor","grade","count"]
    with open(OUTPUT, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"Saved → {OUTPUT}")

if __name__ == "__main__":
    main()
