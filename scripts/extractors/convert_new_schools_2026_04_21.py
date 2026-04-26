#!/usr/bin/env python3
"""
GradeView — New Schools Conversion Script (2026-04-21)
Converts XLSX grade distribution files for three schools to normalized CSV.

Output columns: school_id, year, semester, dept, course_number, instructor, grade, count
  - One row per grade letter per section
  - Rows where count == 0 or null are skipped
  - Non-standard/non-letter grades are skipped
  - Semester normalized to uppercase: FALL, SPRING, SUMMER, WINTER
  - Year as 4-digit string

Schools:
  - ncatsu: North Carolina A&T State University
  - sfasu:  Stephen F. Austin State University
  - unr:    University of Nevada, Reno
"""

import csv
import os
import re
import sys

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
RAW_DIR = os.path.expanduser("~/Documents/GradeView/raw_data")
OUT_DIR  = os.path.expanduser("~/Documents/GradeView/raw_data")

SCHOOLS = {
    "ncatsu": {
        "file": os.path.join(RAW_DIR, "ncatsu", "North Carolina A.T. Grades Data.xlsx"),
        "out":  os.path.join(OUT_DIR, "ncatsu", "ncatsu_converted.csv"),
    },
    "sfasu": {
        "file": os.path.join(RAW_DIR, "sfasu", "StephenFAustin.xlsx"),
        "out":  os.path.join(OUT_DIR, "sfasu", "sfasu_converted.csv"),
    },
    "unr": {
        "file": os.path.join(RAW_DIR, "unr", "University of Nevada, Reno.xlsx"),
        "out":  os.path.join(OUT_DIR, "unr", "unr_converted.csv"),
    },
}

OUTPUT_COLS = ["school_id", "year", "semester", "dept", "course_number", "instructor", "grade", "count"]


# ── Helpers ────────────────────────────────────────────────────────────────────

def parse_semester_str(s):
    """Parse 'Fall 2020' or 'Spring 2021' → (year_str, term_str).
    Returns ('0000', 'UNKNOWN') on failure.
    """
    s = (s or "").strip()
    year_m = re.search(r'\b((?:19|20)\d{2})\b', s)
    year = year_m.group(1) if year_m else "0000"
    term_part = re.sub(r'\b(?:19|20)\d{2}\b', '', s).strip().upper()
    # Normalize known terms
    for canonical in ("FALL", "SPRING", "SUMMER", "WINTER"):
        if canonical in term_part:
            return year, canonical
    return year, term_part or "UNKNOWN"


def safe_int(val):
    """Convert val to int, return 0 on failure or empty."""
    if val is None or str(val).strip() == "":
        return 0
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return 0


# ── NC A&T Extractor ───────────────────────────────────────────────────────────
# Wide format with 2-row header.
# Row 1: metadata column names; grade columns all labelled "Grades" / None
# Row 2: grade letters (A, B, C, D, F, W, I, P, NR, AU, U)
# Data starts row 3.
# Grade cols to keep: A, B, C, D, F, W, I, P  (skip NR, AU, U)

NCATSU_KEEP_GRADES = {"A", "B", "C", "D", "F", "W", "I", "P"}


def extract_ncatsu(filepath, outpath):
    print(f"\n[ncatsu] Reading {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    rows_written = 0
    rows_skipped = 0
    semesters_seen = set()

    with open(outpath, "w", newline="", encoding="utf-8") as fout:
        writer = csv.writer(fout)
        writer.writerow(OUTPUT_COLS)

        row_iter = ws.iter_rows(values_only=True)

        # Row 1: top-level headers
        header1 = list(next(row_iter))
        # Row 2: grade sub-headers (columns 7-17, 0-indexed 6-16)
        header2 = list(next(row_iter))

        # Build effective column index map
        # Columns 0-5: metadata; columns 6-16: grade letters (from row2); col 17: total
        # header2 at index 6..16 = A, B, C, D, F, W, I, P, NR, AU, U
        grade_col_map = {}  # col_index → grade_letter
        for idx in range(6, len(header2)):
            letter = str(header2[idx]).strip() if header2[idx] is not None else None
            if letter and letter in NCATSU_KEEP_GRADES:
                grade_col_map[idx] = letter

        for raw_row in row_iter:
            row = list(raw_row)
            if not any(row):
                rows_skipped += 1
                continue

            snapshot_term = str(row[1]).strip() if row[1] is not None else ""
            dept = str(row[3]).strip() if row[3] is not None else ""       # Course Subject Code
            course_num = str(row[4]).strip() if row[4] is not None else "" # Course Number
            instructor = str(row[5]).strip() if row[5] is not None else "" # Instructor Full Name
            if not instructor:
                instructor = "N/A"

            year, semester = parse_semester_str(snapshot_term)
            semesters_seen.add(f"{semester} {year}")

            if not dept or not course_num or year == "0000":
                rows_skipped += 1
                continue

            for col_idx, grade_letter in grade_col_map.items():
                count = safe_int(row[col_idx] if col_idx < len(row) else None)
                if count <= 0:
                    continue
                writer.writerow([
                    "ncatsu", year, semester, dept, course_num,
                    instructor, grade_letter, count
                ])
                rows_written += 1

    wb.close()
    print(f"[ncatsu] ✅ Done — {rows_written:,} rows written, {rows_skipped:,} skipped")
    print(f"[ncatsu]    Semesters: {sorted(semesters_seen)}")
    return rows_written


# ── SFA Extractor ──────────────────────────────────────────────────────────────
# Standard wide format, 1 header row.
# Grade cols: A, B, C, D, F (skip Passing — developmental only)

SFA_GRADE_COLS = {
    "# Students Earned A": "A",
    "# Students Earned B": "B",
    "# Students Earned C": "C",
    "# Students Earned D": "D",
    "# Students Earned F/Failing": "F",
}


def extract_sfasu(filepath, outpath):
    print(f"\n[sfasu] Reading {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    rows_written = 0
    rows_skipped = 0
    semesters_seen = set()

    row_iter = ws.iter_rows(values_only=True)
    header_raw = list(next(row_iter))
    headers = [str(h).strip() if h is not None else "" for h in header_raw]
    col = {h: i for i, h in enumerate(headers)}

    # Verify expected columns exist
    required = ["Semester", "Year", "Course Prefix", "Course Number",
                "Faculty First Name", "Faculty Last Name"]
    for r in required:
        if r not in col:
            print(f"[sfasu] ❌ Missing expected column: {r}")
            print(f"[sfasu]    Found: {headers}")
            sys.exit(1)

    with open(outpath, "w", newline="", encoding="utf-8") as fout:
        writer = csv.writer(fout)
        writer.writerow(OUTPUT_COLS)

        for raw_row in row_iter:
            row = list(raw_row)
            if not any(row):
                rows_skipped += 1
                continue

            def get(colname):
                idx = col.get(colname)
                return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ""

            semester_word = get("Semester").upper()  # "Spring" → "SPRING"
            year_raw = get("Year")
            # Normalize semester
            for canonical in ("FALL", "SPRING", "SUMMER", "WINTER"):
                if canonical in semester_word:
                    semester_word = canonical
                    break

            year = year_raw[:4] if len(year_raw) >= 4 else year_raw
            if not year or year == "0":
                rows_skipped += 1
                continue

            dept = get("Course Prefix")
            course_num = get("Course Number")
            first = get("Faculty First Name")
            last = get("Faculty Last Name")
            instructor = f"{first} {last}".strip() or "N/A"

            semesters_seen.add(f"{semester_word} {year}")

            if not dept or not course_num:
                rows_skipped += 1
                continue

            for col_name, grade_letter in SFA_GRADE_COLS.items():
                count = safe_int(row[col[col_name]] if col_name in col and col[col_name] < len(row) else None)
                if count <= 0:
                    continue
                writer.writerow([
                    "sfasu", year, semester_word, dept, course_num,
                    instructor, grade_letter, count
                ])
                rows_written += 1

    wb.close()
    print(f"[sfasu] ✅ Done — {rows_written:,} rows written, {rows_skipped:,} skipped")
    print(f"[sfasu]    Semesters: {sorted(semesters_seen)}")
    return rows_written


# ── UNR Extractor ──────────────────────────────────────────────────────────────
# Wide format with plus/minus grades, 1 header row.
# Grade cols to keep: A, A-, B+, B, B-, C+, C, C-, D+, D, D-, F, W
# Skip: S, U, AD, I, R, X

UNR_KEEP_GRADES = {"A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F", "W"}


def extract_unr(filepath, outpath):
    print(f"\n[unr] Reading {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    rows_written = 0
    rows_skipped = 0
    semesters_seen = set()

    row_iter = ws.iter_rows(values_only=True)
    header_raw = list(next(row_iter))
    # Header has a None in col 4 (blank) — keep as empty string
    headers = [str(h).strip() if h is not None else "" for h in header_raw]

    # Build grade col indices: columns whose headers are in UNR_KEEP_GRADES
    grade_col_indices = []
    for idx, h in enumerate(headers):
        if h in UNR_KEEP_GRADES:
            grade_col_indices.append((idx, h))

    col = {h: i for i, h in enumerate(headers) if h}

    required = ["Term", "Prefix", "Number", "Instructor"]
    for r in required:
        if r not in col:
            print(f"[unr] ❌ Missing expected column: {r}")
            print(f"[unr]    Found: {headers}")
            sys.exit(1)

    with open(outpath, "w", newline="", encoding="utf-8") as fout:
        writer = csv.writer(fout)
        writer.writerow(OUTPUT_COLS)

        for raw_row in row_iter:
            row = list(raw_row)
            if not any(row):
                rows_skipped += 1
                continue

            def get(colname):
                idx = col.get(colname)
                return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ""

            term_str = get("Term")          # e.g. "Fall 2020"
            dept = get("Prefix")            # subject prefix
            course_num = get("Number")      # course number
            instructor = get("Instructor")  # "Lastname,Firstname M."
            if not instructor:
                instructor = "N/A"

            year, semester = parse_semester_str(term_str)
            semesters_seen.add(f"{semester} {year}")

            if not dept or not course_num or year == "0000":
                rows_skipped += 1
                continue

            for col_idx, grade_letter in grade_col_indices:
                count = safe_int(row[col_idx] if col_idx < len(row) else None)
                if count <= 0:
                    continue
                writer.writerow([
                    "unr", year, semester, dept, course_num,
                    instructor, grade_letter, count
                ])
                rows_written += 1

    wb.close()
    print(f"[unr] ✅ Done — {rows_written:,} rows written, {rows_skipped:,} skipped")
    print(f"[unr]    Semesters: {sorted(semesters_seen)}")
    return rows_written


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    results = {}
    for school_id, paths in SCHOOLS.items():
        filepath = paths["file"]
        outpath = paths["out"]
        if not os.path.exists(filepath):
            print(f"❌ File not found: {filepath}")
            sys.exit(1)

        if school_id == "ncatsu":
            count = extract_ncatsu(filepath, outpath)
        elif school_id == "sfasu":
            count = extract_sfasu(filepath, outpath)
        elif school_id == "unr":
            count = extract_unr(filepath, outpath)
        else:
            print(f"Unknown school: {school_id}")
            sys.exit(1)

        results[school_id] = {"count": count, "output": outpath}
        print(f"   → Output: {outpath}")

    print("\n" + "=" * 60)
    print("CONVERSION SUMMARY")
    print("=" * 60)
    for school_id, info in results.items():
        print(f"  {school_id}: {info['count']:,} rows → {info['output']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
