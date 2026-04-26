#!/usr/bin/env python3
"""NIU extractor — wide format, grades as columns. Term: 'Fall 2020', Class: 'ACCY 331 00H1'."""
import csv, glob, re
from pathlib import Path
import openpyxl

RAW_DIR = Path.home() / "Documents/GradeView/raw_data/northern_illinois"
OUTPUT  = RAW_DIR / "niu_grades_extracted.csv"

GRADE_COLS = ['A','A-','B+','B','B-','C+','C','C-','D','F','P','W']

SEMESTER_MAP = {'fall':'FALL','spring':'SPRING','summer':'SUMMER','winter':'WINTER'}

def parse_term(term):
    """'Fall 2020' → (2020, FALL)"""
    m = re.match(r'(\w+)\s+(\d{4})', str(term).strip())
    if m:
        sem = SEMESTER_MAP.get(m.group(1).lower(), m.group(1).upper())
        return m.group(2), sem
    return '0000', 'UNKNOWN'

def parse_class(cls):
    """'ACCY 331 00H1' → (dept='ACCY', course='331')"""
    parts = str(cls).strip().split()
    dept   = parts[0] if len(parts) > 0 else ''
    course = parts[1] if len(parts) > 1 else ''
    return dept, course

def main():
    files = sorted(glob.glob(str(RAW_DIR / "*.xlsx")))
    all_rows = []
    for f in files:
        print(f"  {Path(f).name}...")
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows: continue
            headers = [str(h).strip() if h else '' for h in rows[0]]
            try:
                i_term  = headers.index('Term')
                i_class = headers.index('Class and Section')
                i_instr = headers.index('Primary Instructor')
            except ValueError:
                continue
            grade_idx = {h: i for i,h in enumerate(headers) if h in GRADE_COLS}
            for row in rows[1:]:
                if not row or not row[i_term]: continue
                year, semester = parse_term(row[i_term])
                dept, course = parse_class(row[i_class])
                instr = str(row[i_instr]).strip() if row[i_instr] else 'N/A'
                # Format "Riley,Mark Edward" → "Riley, Mark Edward"
                if ',' in instr and ', ' not in instr:
                    instr = instr.replace(',', ', ', 1)
                for grade, col_i in grade_idx.items():
                    count = row[col_i]
                    try: count = int(count)
                    except: count = 0
                    if count <= 0: continue
                    all_rows.append({
                        'school_id':'niu','year':year,'semester':semester,
                        'dept':dept,'course_number':course,'instructor':instr,
                        'grade':grade,'count':count
                    })
        wb.close()
        print(f"    → {len(all_rows):,} rows so far")

    print(f"\nTotal: {len(all_rows):,}")
    fieldnames = ['school_id','year','semester','dept','course_number','instructor','grade','count']
    with open(OUTPUT, 'w', newline='', encoding='utf-8') as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)
    print(f"Saved → {OUTPUT}")

if __name__ == '__main__':
    main()
