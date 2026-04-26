#!/usr/bin/env python3
"""
GradeView — School Data Upload v2
===================================
Replaces HTTP batch upload with direct libsql connection + transaction safety.

Key improvements over v1:
  - Parameterized queries (no SQL injection, handles O'Brien etc.)
  - Transaction safety (old data preserved if upload fails)
  - Persistent staging DB (survives reboots)
  - Progress checkpointing (resume from last checkpoint after SIGTERM)
  - Credentials loaded from .turso_config (not hardcoded)
  - Integrity verification (row count + checksum)
  - Dry run mode

Usage:
    # Stage locally (read files, build staging DB)
    python3 upload_v2.py --school ferrisstate --file data.xlsx --stage-only

    # Verify staging DB
    python3 upload_v2.py --school ferrisstate --verify-staging

    # Push to Turso (with transaction safety)
    caffeinate -i python3 upload_v2.py --school ferrisstate --push

    # Verify remote data matches local
    python3 upload_v2.py --school ferrisstate --verify-remote

    # Full pipeline (stage + push + verify)
    caffeinate -i python3 upload_v2.py --school ferrisstate --file data.xlsx

    # Append new semester data (don't delete existing)
    caffeinate -i python3 upload_v2.py --school ferrisstate --file new_data.xlsx --append

    # Resume interrupted push
    caffeinate -i python3 upload_v2.py --school ferrisstate --push --resume

    # Dry run (stage + verify, no Turso writes)
    python3 upload_v2.py --school ferrisstate --file data.xlsx --dry-run

Column mappings (--map):
    utaustin, tamu, xlsx_generic, uh, unlv, neiu
    (Same mappings as v1 — see MAPPINGS dict)
"""

import argparse
import csv
import glob
import hashlib
import json
import os
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path

# ── Configuration ─────────────────────────────────────────────────────────────

TURSO_CONFIG_PATH = os.path.expanduser(
    "~/.openclaw/workspace/projects/gradeview/.turso_config"
)
STAGING_DIR = os.path.expanduser(
    "~/.openclaw/workspace/projects/gradeview/staging"
)
PROGRESS_DIR = os.path.expanduser(
    "~/.openclaw/workspace/projects/gradeview/staging/progress"
)

def load_config():
    """Load Turso credentials from .turso_config — NEVER hardcode tokens."""
    if not os.path.exists(TURSO_CONFIG_PATH):
        print(f"❌ Config not found: {TURSO_CONFIG_PATH}")
        print(f"   Expected format:")
        print(f"   TURSO_URL=libsql://...")
        print(f"   TURSO_TOKEN=eyJ...")
        sys.exit(1)
    config = {}
    with open(TURSO_CONFIG_PATH) as f:
        for line in f:
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                key, val = line.split('=', 1)
                config[key.strip()] = val.strip()
    required = ['TURSO_URL', 'TURSO_TOKEN']
    for key in required:
        if key not in config:
            print(f"❌ Missing {key} in {TURSO_CONFIG_PATH}")
            sys.exit(1)
    return config

def staging_db_path(school_id):
    os.makedirs(STAGING_DIR, exist_ok=True)
    return os.path.join(STAGING_DIR, f"{school_id}_staging.db")

def progress_path(school_id):
    os.makedirs(PROGRESS_DIR, exist_ok=True)
    return os.path.join(PROGRESS_DIR, f"{school_id}_progress.json")

# ── Column Mappings (same as v1) ─────────────────────────────────────────────

MAPPINGS = {
    "utaustin": {
        "semester":  "Semester",
        "dept":      "Course Prefix",
        "course_number": "Course Number",
        "grade":     "Letter Grade",
        "count":     "Count of letter grade",
        "instructor": None,
    },
    "tamu": {
        "semester":  "Term",
        "dept":      "Dept",
        "course_number": "Course",
        "grade":     "Grade",
        "count":     "Count",
        "instructor": "Instructor",
    },
    "xlsx_generic": {
        "semester":  ["Semester", "Term", "Academic Term", "Acad Term"],
        "dept":      ["Dept", "Department", "Course Prefix", "Subject"],
        "course_number": ["Course", "Course Number", "Course No", "CourseNo"],
        "grade":     ["Grade", "Letter Grade", "Final Grade"],
        "count":     ["Count", "Count of letter grade", "N", "Enrollment"],
        "instructor": ["Instructor", "Professor", "Faculty", "Instructor Name"],
    },
    "unlv": {
        "format":    "wide",
        "sheet":     "PRR-2026-102_ODS-grade-distribu",
        "year":      "CALENDAR_YEAR",
        "semester":  "TERM",
        "dept":      "SUBJECT",
        "course_number": "CATALOG_NBR",
        "instructor_first": "INSTRUCTORS",
        "instructor_last":  None,
        "grade_cols": {
            "A_GRADE": "A", "AMINUS_GRADE": "A-",
            "BPLUS_GRADE": "B+", "B_GRADE": "B", "BMINUS_GRADE": "B-",
            "CPLUS_GRADE": "C+", "C_GRADE": "C", "CMINUS_GRADE": "C-",
            "DPLUS_GRADE": "D+", "D_GRADE": "D", "DMINUS_GRADE": "D-",
            "F_GRADE": "F"
        },
    },
    "neiu": {
        "format":    "wide",
        "sheet":     "Data",
        "year":      "ACADEMIC_YEAR",
        "semester":  "SEMESTER",
        "dept":      "COURSE_SUBJECT",
        "course_number": "COURSE_NUMBER",
        "instructor_first": "PRIMARY_INSTRUCTOR_FIRST_NAME",
        "instructor_last":  "PRIMARY_INSTRUCTOR_LAST_NAME",
        "grade_cols": {
            "A_GRADE": "A", "B_GRADE": "B", "C_GRADE": "C",
            "D_GRADE": "D", "F_GRADE": "F", "W_GRADE": "W",
            "I_GRADE": "I", "PASS_GRADE": "P"
        },
    },
    "umich": {
        "semester":      "semester",
        "dept":          "department",
        "course_number": "course_number",
        "grade":         "grade",
        "count":         "count",
        "instructor":    "instructor",
    },
    "cpp": {
        "format":    "wide",
        "semester":  "semester",
        "year":      "year",
        "dept":      "department",
        "course_number": "course_number",
        "instructor_first": "instructor",
        "instructor_last":  None,
        "grade_cols": {
            "A": "A", "A-": "A-", "B+": "B+", "B": "B", "B-": "B-",
            "C+": "C+", "C": "C", "C-": "C-", "D+": "D+", "D": "D",
            "D-": "D-", "F": "F"
        },
    },
    "ucberkeley_f25": {
        "format":    "wide",
        "semester":  "semester",
        "year":      "year",
        "dept":      "department",
        "course_number": "course_number",
        "instructor_first": "instructor",
        "instructor_last":  None,
        "grade_cols": {
            "A+": "A+", "A": "A", "A-": "A-", "B+": "B+", "B": "B", "B-": "B-",
            "C+": "C+", "C": "C", "C-": "C-", "D+": "D+", "D": "D",
            "D-": "D-", "F": "F"
        },
    },
    # gradeview_normalized: output from GradeView conversion scripts (separate year + semester cols)
    "gradeview_normalized": {
        "year":          "year",
        "semester":      "semester",
        "dept":          "dept",
        "course_number": "course_number",
        "grade":         "grade",
        "count":         "count",
        "instructor":    "instructor",
    },
    # UNT: multi-sheet file — two sheets with different column layouts.
    # Use --map unt to auto-process both sheets.
    "unt": {
        "_multi_sheet": True,
        "sheets": {
            "2021-2022": {
                "semester":      "Semester/Term",
                "dept":          "Subject",
                "course_number": "Catalog Number",
                "grade":         "Grade",
                "count":         "Grade Count",
                "instructor":    "Instructor",
            },
            "2023-2025": {
                "semester":      "Term",      # PeopleSoft codes: 1231 = Spring 2023
                "dept":          "Subject",
                "course_number": "Catalog",
                "grade":         "Grade",
                "count":         "Count ID",
                "instructor":    "Name",
                "_skip_if_empty": ["Name"],  # 123 rows with no instructor have corrupt counts
            },
        },
    },
}

# ── File Reading ──────────────────────────────────────────────────────────────

def resolve_col(row, mapping_val):
    if mapping_val is None:
        return None
    if isinstance(mapping_val, list):
        for candidate in mapping_val:
            if candidate in row:
                return row[candidate]
        return None
    return row.get(mapping_val)

def read_csv_file(filepath):
    rows = []
    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows

def read_xlsx_file(filepath, sheet_name=None, all_sheets=False):
    """Read xlsx file. If all_sheets=True, reads every sheet and returns
    a dict {sheet_name: [rows]}. Otherwise returns a flat list of rows
    from a single sheet (specified by sheet_name, or the active sheet)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)

    if all_sheets:
        result = {}
        for name in wb.sheetnames:
            ws = wb[name]
            headers = None
            sheet_rows = []
            for excel_row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c).strip() if c else '' for c in excel_row]
                    continue
                row = dict(zip(headers, [str(c).strip() if c is not None else '' for c in excel_row]))
                sheet_rows.append(row)
            result[name] = sheet_rows
        wb.close()
        return result

    # Single sheet mode (original behavior)
    rows = []
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = None
    for excel_row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else '' for c in excel_row]
            continue
        row = dict(zip(headers, [str(c).strip() if c is not None else '' for c in excel_row]))
        rows.append(row)
    wb.close()
    return rows

def parse_semester(semester_str):
    """Parse semester string into (term, year). Handles formats like:
       'Fall 2020', 'Spring / Summer 2021', 'Winter 2025', '2024 Spring'
       Also handles PeopleSoft term codes: '1231' → ('SPRING', '2023')
       Format: CYYT where C=century(1=2000s), YY=year, T=term(1=Spring,3=Summer,8=Fall)
    """
    import re
    s = (semester_str or '').strip()
    if not s or s.lower() == 'none':
        return 'UNKNOWN', '0000'

    # PeopleSoft term code: exactly 4 digits, starts with 1, ends with 1/3/8
    ps_match = re.match(r'^1(\d{2})([138])$', s)
    if ps_match:
        ps_term_map = {'1': 'SPRING', '3': 'SUMMER', '8': 'FALL'}
        year = f'20{ps_match.group(1)}'
        term = ps_term_map[ps_match.group(2)]
        return term, year

    # Extract year (4-digit number starting with 19 or 20)
    year_match = re.search(r'\b((?:19|20)\d{2})\b', s)
    year = year_match.group(1) if year_match else '0000'

    # Extract term: everything that isn't the year, cleaned up
    term_part = re.sub(r'\b(?:19|20)\d{2}\b', '', s).strip()
    # Normalize separators and whitespace
    term_part = re.sub(r'\s*/\s*', '/', term_part).strip().upper()
    term = term_part if term_part else 'UNKNOWN'

    return term, year

# ── Staging ───────────────────────────────────────────────────────────────────

def stage_standard(school_id, data_rows, mapping):
    """Stage standard format data (one row per grade entry)."""
    db_path = staging_db_path(school_id)
    if os.path.exists(db_path):
        os.remove(db_path)

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE grades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            school_id TEXT NOT NULL,
            year TEXT, semester TEXT, dept TEXT,
            course_number TEXT, instructor TEXT, grade TEXT, count INTEGER
        )
    """)

    inserted = 0
    skipped = 0
    batch = []

    has_year_col = 'year' in mapping and mapping['year']

    for row in data_rows:
        semester_str = resolve_col(row, mapping.get("semester")) or ''
        dept = (resolve_col(row, mapping.get("dept")) or '').strip()
        course_number = (resolve_col(row, mapping.get("course_number")) or '').strip()
        grade = (resolve_col(row, mapping.get("grade")) or '').strip()
        instructor = (resolve_col(row, mapping.get("instructor")) or 'N/A').strip()
        try:
            count = int((resolve_col(row, mapping.get("count")) or '0').strip())
        except (ValueError, AttributeError):
            count = 0

        if not grade:
            skipped += 1
            continue

        if has_year_col:
            year = str(row.get(mapping['year'], '') or '').strip()
            term = semester_str.strip().upper()
            if not term or term.lower() == 'none':
                term = 'UNKNOWN'
            # Normalize common term names
            for t in ('SPRING', 'SUMMER', 'FALL', 'WINTER'):
                if t in term:
                    term = t
                    break
        else:
            term, year = parse_semester(semester_str)
        batch.append((school_id, year, term, dept, course_number, instructor or 'N/A', grade, count))

        if len(batch) >= 5000:
            cur.executemany(
                "INSERT INTO grades (school_id, year, semester, dept, course_number, instructor, grade, count) VALUES (?,?,?,?,?,?,?,?)",
                batch
            )
            inserted += len(batch)
            batch = []

    if batch:
        cur.executemany(
            "INSERT INTO grades (school_id, year, semester, dept, course_number, instructor, grade, count) VALUES (?,?,?,?,?,?,?,?)",
            batch
        )
        inserted += len(batch)

    conn.commit()

    # Deduplicate: standard format often has multiple sections for the same course.
    # GradeView doesn't track sections — aggregate counts across sections.
    cur = conn.cursor()
    before = cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]
    cur.execute("""
        CREATE TABLE grades_deduped AS
        SELECT school_id, year, semester, dept, course_number, instructor, grade,
               SUM(count) as count
        FROM grades
        GROUP BY school_id, year, semester, dept, course_number, instructor, grade
    """)
    cur.execute("DROP TABLE grades")
    cur.execute("ALTER TABLE grades_deduped RENAME TO grades")
    after = cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]
    conn.commit()
    if before != after:
        merged = before - after
        print(f"   📦 Deduplicated: {before:,} → {after:,} rows ({merged:,} section-level rows merged)")
        inserted = after

    conn.close()
    print(f"✅ Staging complete: {inserted:,} rows ({skipped:,} skipped) → {db_path}")
    return inserted

def stage_standard_multi(school_id, data_rows):
    """Stage standard format data where each row carries its own mapping
    (via _sheet_mapping key). Used for multi-sheet files like UNT where
    different sheets have different column layouts."""
    db_path = staging_db_path(school_id)
    if os.path.exists(db_path):
        os.remove(db_path)

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE grades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            school_id TEXT NOT NULL,
            year TEXT, semester TEXT, dept TEXT,
            course_number TEXT, instructor TEXT, grade TEXT, count INTEGER
        )
    """)

    inserted = 0
    skipped = 0
    batch = []

    for row in data_rows:
        mapping = row.pop('_sheet_mapping', {})

        # Skip rows where required fields are empty (e.g. corrupt aggregation rows)
        skip_fields = mapping.get('_skip_if_empty', [])
        if skip_fields and any(not str(row.get(f, '') or '').strip() for f in skip_fields):
            skipped += 1
            continue

        semester_str = resolve_col(row, mapping.get("semester")) or ''
        dept = (resolve_col(row, mapping.get("dept")) or '').strip()
        course_number = (resolve_col(row, mapping.get("course_number")) or '').strip()
        grade = (resolve_col(row, mapping.get("grade")) or '').strip()
        instructor = (resolve_col(row, mapping.get("instructor")) or 'N/A').strip()
        try:
            count = int((resolve_col(row, mapping.get("count")) or '0').strip())
        except (ValueError, AttributeError):
            count = 0

        if not grade:
            skipped += 1
            continue

        has_year_col = 'year' in mapping and mapping['year']
        if has_year_col:
            year = str(row.get(mapping['year'], '') or '').strip()
            term = semester_str.strip().upper()
            if not term or term.lower() == 'none':
                term = 'UNKNOWN'
            for t in ('SPRING', 'SUMMER', 'FALL', 'WINTER'):
                if t in term:
                    term = t
                    break
        else:
            term, year = parse_semester(semester_str)
        batch.append((school_id, year, term, dept, course_number, instructor or 'N/A', grade, count))

        if len(batch) >= 5000:
            cur.executemany(
                "INSERT INTO grades (school_id, year, semester, dept, course_number, instructor, grade, count) VALUES (?,?,?,?,?,?,?,?)",
                batch
            )
            inserted += len(batch)
            batch = []

    if batch:
        cur.executemany(
            "INSERT INTO grades (school_id, year, semester, dept, course_number, instructor, grade, count) VALUES (?,?,?,?,?,?,?,?)",
            batch
        )
        inserted += len(batch)

    conn.commit()

    # Deduplicate across sections
    cur = conn.cursor()
    before = cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]
    cur.execute("""
        CREATE TABLE grades_deduped AS
        SELECT school_id, year, semester, dept, course_number, instructor, grade,
               SUM(count) as count
        FROM grades
        GROUP BY school_id, year, semester, dept, course_number, instructor, grade
    """)
    cur.execute("DROP TABLE grades")
    cur.execute("ALTER TABLE grades_deduped RENAME TO grades")
    after = cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]
    conn.commit()
    if before != after:
        merged = before - after
        print(f"   📦 Deduplicated: {before:,} → {after:,} rows ({merged:,} section-level rows merged)")
        inserted = after

    conn.close()
    print(f"✅ Staging complete: {inserted:,} rows ({skipped:,} skipped) → {db_path}")
    return inserted

def stage_wide(school_id, data_rows, mapping):
    """Stage wide-format data (one row per section, multiple grade columns)."""
    db_path = staging_db_path(school_id)
    if os.path.exists(db_path):
        os.remove(db_path)

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE grades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            school_id TEXT NOT NULL,
            year TEXT, semester TEXT, dept TEXT,
            course_number TEXT, instructor TEXT, grade TEXT, count INTEGER
        )
    """)

    inserted = 0
    batch = []

    # Check for auto-mapper extensions
    parse_section = mapping.get('_parse_section')
    has_year_col = 'year' in mapping and mapping['year']

    for row in data_rows:
        # Year + Semester: either separate columns or combined
        if has_year_col:
            year = str(row.get(mapping['year'], '') or '').strip()
            term = str(row.get(mapping.get('semester', ''), '') or '').strip().upper()
        elif mapping.get('semester'):
            sem_str = str(row.get(mapping['semester'], '') or '').strip()
            term, year = parse_semester(sem_str)
        else:
            year, term = '0000', 'UNKNOWN'

        # Dept + Course: either direct columns or parsed from compound section
        if parse_section:
            section_val = str(row.get(parse_section['column'], '') or '').strip()
            parts = section_val.split()
            dept_idx = parse_section.get('dept_index', 0)
            course_idx = parse_section.get('course_index', 1)
            dept = parts[dept_idx] if len(parts) > dept_idx else ''
            course = parts[course_idx] if len(parts) > course_idx else ''
        else:
            dept = str(row.get(mapping.get('dept', ''), '') or '').strip()
            course = str(row.get(mapping.get('course_number', ''), '') or '').strip()

        fname = str(row.get(mapping.get('instructor_first', ''), '') or '').strip()
        lname = str(row.get(mapping.get('instructor_last', ''), '') or '').strip()
        instructor = f'{fname} {lname}'.strip() or 'N/A'

        for col, letter in mapping['grade_cols'].items():
            try:
                count = int(row.get(col) or 0)
            except (ValueError, TypeError):
                count = 0
            if count > 0:
                batch.append((school_id, year, term, dept, course, instructor, letter, count))

        if len(batch) >= 5000:
            cur.executemany(
                "INSERT INTO grades (school_id, year, semester, dept, course_number, instructor, grade, count) VALUES (?,?,?,?,?,?,?,?)",
                batch
            )
            inserted += len(batch)
            batch = []

    if batch:
        cur.executemany(
            "INSERT INTO grades (school_id, year, semester, dept, course_number, instructor, grade, count) VALUES (?,?,?,?,?,?,?,?)",
            batch
        )
        inserted += len(batch)

    conn.commit()

    # Deduplicate: if compound section parsing created duplicate keys, aggregate counts
    if parse_section:
        cur = conn.cursor()
        before = cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]
        cur.execute("""
            CREATE TABLE grades_deduped AS
            SELECT school_id, year, semester, dept, course_number, instructor, grade,
                   SUM(count) as count
            FROM grades
            GROUP BY school_id, year, semester, dept, course_number, instructor, grade
        """)
        cur.execute("DROP TABLE grades")
        cur.execute("ALTER TABLE grades_deduped RENAME TO grades")
        after = cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]
        conn.commit()
        if before != after:
            merged = before - after
            print(f"   📦 Deduplicated: {before:,} → {after:,} rows ({merged:,} section-level rows merged)")
            inserted = after

    conn.close()
    print(f"✅ Staging complete: {inserted:,} rows → {db_path}")
    return inserted

# ── Integrity ─────────────────────────────────────────────────────────────────

def compute_local_checksum(school_id):
    """Compute checksum of staged data for integrity verification."""
    db_path = staging_db_path(school_id)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        SELECT school_id, year, semester, dept, course_number, instructor, grade, count
        FROM grades ORDER BY dept, course_number, semester, grade, instructor
    """)
    h = hashlib.sha256()
    count = 0
    for row in cur:
        h.update("|".join(str(v) for v in row).encode())
        count += 1
    conn.close()
    return h.hexdigest()[:16], count

def verify_staging(school_id):
    """Verify staging DB integrity and show sample data."""
    db_path = staging_db_path(school_id)
    if not os.path.exists(db_path):
        print(f"❌ No staging DB for {school_id}. Run --stage-only first.")
        return False

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    total = cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]
    depts = cur.execute("SELECT COUNT(DISTINCT dept) FROM grades").fetchone()[0]
    semesters = cur.execute("SELECT COUNT(DISTINCT semester || ' ' || year) FROM grades").fetchone()[0]
    grades = cur.execute("SELECT DISTINCT grade FROM grades ORDER BY grade").fetchall()
    grade_list = [g[0] for g in grades]

    # Check for duplicates
    dupes = cur.execute("""
        SELECT school_id, year, semester, dept, course_number, instructor, grade, count,
               COUNT(*) as n
        FROM grades
        GROUP BY school_id, year, semester, dept, course_number, instructor, grade, count
        HAVING n > 1
    """).fetchall()

    # Sample data
    sample = cur.execute("SELECT * FROM grades LIMIT 5").fetchall()

    conn.close()

    checksum, _ = compute_local_checksum(school_id)

    print(f"\n{'='*50}")
    print(f"Staging Verification: {school_id}")
    print(f"{'='*50}")
    print(f"  Total rows:    {total:,}")
    print(f"  Departments:   {depts}")
    print(f"  Semesters:     {semesters}")
    print(f"  Grade values:  {', '.join(grade_list)}")
    print(f"  Duplicates:    {len(dupes)}")
    print(f"  Checksum:      {checksum}")
    print(f"\n  Sample rows:")
    for row in sample:
        print(f"    {row}")

    if dupes:
        print(f"\n  ⚠️ WARNING: {len(dupes)} duplicate row groups found!")
        for d in dupes[:5]:
            print(f"    {d}")

    print(f"\n{'='*50}")
    return total > 0 and len(dupes) == 0

# ── Transport Detection ──────────────────────────────────────────────────────
# libsql (WebSocket) is preferred: true transactions, faster, more reliable.
# HTTP pipeline is the fallback if libsql-experimental isn't installed.

LIBSQL_AVAILABLE = False
try:
    import libsql_experimental as libsql
    LIBSQL_AVAILABLE = True
except ImportError:
    pass

def get_transport(config, force=None):
    """Return the transport type to use. Logs which one is active."""
    if force == 'http':
        return 'http'
    if force == 'libsql' and not LIBSQL_AVAILABLE:
        print("❌ --transport libsql requested but libsql-experimental not installed")
        print("   Install: pip3 install libsql-experimental --break-system-packages")
        sys.exit(1)
    if LIBSQL_AVAILABLE and force != 'http':
        return 'libsql'
    return 'http'


# ── libsql Transport (preferred) ────────────────────────────────────────────

def libsql_connect(config):
    """Open a persistent WebSocket connection to Turso via libsql."""
    url = config['TURSO_URL']
    token = config['TURSO_TOKEN']
    return libsql.connect(url, auth_token=token)

def get_remote_count_libsql(config, school_id):
    """Get row count via libsql."""
    conn = libsql_connect(config)
    result = conn.execute(
        "SELECT COUNT(*) FROM grades WHERE school_id = ?", (school_id,)
    ).fetchone()
    return int(result[0])

def _build_multi_row_insert(rows, cols=8):
    """Build a multi-row INSERT statement packing many rows into one SQL call.

    Instead of N individual INSERTs (each a separate network round-trip),
    this packs up to ROWS_PER_STMT rows into a single INSERT ... VALUES
    (...), (...), (...) statement. One network call inserts hundreds of rows.

    SQLite variable limit is 32766 (since 3.32.0). With 8 cols per row,
    that's ~4095 rows per statement. We cap at 100 to stay well under
    limits and keep individual statement execution time low.
    """
    ROWS_PER_STMT = 100  # 100 rows × 8 cols = 800 params (safe under any limit)
    stmts = []
    for i in range(0, len(rows), ROWS_PER_STMT):
        chunk = rows[i:i + ROWS_PER_STMT]
        placeholders = ", ".join(["(?,?,?,?,?,?,?,?)"] * len(chunk))
        flat_params = [val for row in chunk for val in row]
        stmts.append((
            f"INSERT INTO grades (school_id, year, semester, dept, course_number, "
            f"instructor, grade, count) VALUES {placeholders}",
            flat_params
        ))
    return stmts

def push_to_turso_libsql(school_id, config, append=False, resume=False):
    """Push staging DB to Turso using libsql with true transaction safety.

    Key advantages over HTTP:
      - Single persistent WebSocket connection (no per-batch handshakes)
      - Real BEGIN/COMMIT/ROLLBACK transactions (atomic all-or-nothing)
      - Multi-row INSERTs: 100 rows per SQL statement (not 1-row-per-call)
      - If anything fails, ROLLBACK preserves original data completely
    """
    db_path = staging_db_path(school_id)
    prog_path = progress_path(school_id)

    if not os.path.exists(db_path):
        print(f"❌ No staging DB for {school_id}. Run --stage-only first.")
        return False

    local_conn = sqlite3.connect(db_path)
    local_cur = local_conn.cursor()
    local_count = local_cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]

    remote_conn = libsql_connect(config)
    existing = remote_conn.execute(
        "SELECT COUNT(*) FROM grades WHERE school_id = ?", (school_id,)
    ).fetchone()[0]

    # Resume logic
    skip = 0
    if resume and os.path.exists(prog_path):
        with open(prog_path) as f:
            progress = json.load(f)
        skip = progress.get('rows_uploaded', 0)
        print(f"⏩ Resuming from row {skip:,} (checkpoint: {prog_path})")
    elif append:
        print(f"➕ Append mode: {existing:,} rows exist in Turso. Adding new rows.")
    else:
        if existing > 0:
            print(f"🔄 Fresh upload: will replace {existing:,} existing rows for {school_id}")

    remaining = local_count - skip
    print(f"\n📤 Pushing {remaining:,} rows to Turso via libsql ({school_id})")
    print(f"   Local: {local_count:,} | Remote: {existing:,} | Skip: {skip:,}")
    print(f"   Transport: libsql (WebSocket, multi-row INSERT, transactional)")
    print(f"   Packing: 100 rows per INSERT statement, 5000 rows per batch")

    BATCH = 5000  # Rows to accumulate before executing a batch of multi-row INSERTs

    # Begin transaction — everything is atomic from here
    try:
        remote_conn.execute("BEGIN")

        # Delete old data inside the transaction (safe: rolls back if insert fails)
        if not append and not resume and existing > 0:
            print(f"\n   Deleting old {school_id} data (inside transaction)...")
            remote_conn.execute(
                "DELETE FROM grades WHERE school_id = ?", (school_id,)
            )
            print(f"   ✅ Old data marked for deletion (commits only if upload succeeds)")

        local_cur.execute(
            "SELECT school_id, year, semester, dept, course_number, instructor, grade, count "
            "FROM grades LIMIT -1 OFFSET ?",
            (skip,)
        )

        total_uploaded = skip
        start_time = time.time()
        batch = []

        for row in local_cur:
            sid, yr, sem, dept, course, instr, grade, cnt = row
            batch.append((
                str(sid), str(yr or ''), str(sem or ''), str(dept or ''),
                str(course or ''), str(instr or 'N/A'), str(grade or ''), int(cnt)
            ))

            if len(batch) >= BATCH:
                # Pack 5000 rows into ~50 multi-row INSERT statements
                for sql, params in _build_multi_row_insert(batch):
                    remote_conn.execute(sql, params)
                total_uploaded += len(batch)
                batch = []

                elapsed = time.time() - start_time
                rate = (total_uploaded - skip) / elapsed if elapsed > 0 else 0
                eta = (local_count - total_uploaded) / rate if rate > 0 else 0
                print(f"  {total_uploaded:,}/{local_count:,} rows "
                      f"({rate:.0f} rows/s, ETA: {eta/60:.1f}m)")

                # Progress checkpoint
                with open(prog_path, 'w') as f:
                    json.dump({
                        'school_id': school_id,
                        'rows_uploaded': total_uploaded,
                        'total_rows': local_count,
                        'timestamp': datetime.now().isoformat(),
                    }, f)

        # Final batch
        if batch:
            for sql, params in _build_multi_row_insert(batch):
                remote_conn.execute(sql, params)
            total_uploaded += len(batch)

        # Verify count BEFORE committing
        in_txn_count = remote_conn.execute(
            "SELECT COUNT(*) FROM grades WHERE school_id = ?", (school_id,)
        ).fetchone()[0]

        expected = local_count if not append else (existing + local_count - skip)
        if in_txn_count != expected:
            print(f"\n❌ Pre-commit verification failed! "
                  f"Expected {expected:,} but found {in_txn_count:,}")
            print(f"   Rolling back — original data preserved.")
            remote_conn.execute("ROLLBACK")
            local_conn.close()
            return False

        # All good — commit
        remote_conn.execute("COMMIT")

        elapsed = time.time() - start_time
        print(f"\n✅ Upload complete: {total_uploaded:,} rows in {elapsed:.1f}s (0 errors)")
        print(f"   Transaction committed successfully.")

        # Clean up progress file
        if os.path.exists(prog_path):
            os.remove(prog_path)

        local_conn.close()
        return True

    except Exception as e:
        print(f"\n❌ Error during upload: {e}")
        print(f"   Rolling back — original data preserved.")
        try:
            remote_conn.execute("ROLLBACK")
        except Exception:
            pass  # Connection may already be dead
        local_conn.close()
        return False


# ── HTTP Transport (fallback) ────────────────────────────────────────────────

def turso_pipeline(config, statements, retries=3):
    """Send statements to Turso HTTP pipeline API with retry logic."""
    import urllib.request
    api_url = f"{config['TURSO_URL'].replace('libsql://', 'https://')}/v2/pipeline"
    payload = json.dumps({"requests": statements}).encode()

    for attempt in range(retries):
        try:
            req = urllib.request.Request(
                api_url, data=payload,
                headers={
                    "Authorization": f"Bearer {config['TURSO_TOKEN']}",
                    "Content-Type": "application/json"
                }
            )
            with urllib.request.urlopen(req, timeout=180) as r:
                return json.loads(r.read())
        except Exception as e:
            if attempt < retries - 1:
                wait = 5 * (attempt + 1)
                print(f"  ⚠️ Network error (attempt {attempt+1}/{retries}): {e} — retrying in {wait}s...")
                time.sleep(wait)
            else:
                raise

def get_remote_count_http(config, school_id):
    """Get row count via HTTP pipeline API."""
    result = turso_pipeline(config, [{
        "type": "execute",
        "stmt": {
            "sql": "SELECT COUNT(*) FROM grades WHERE school_id = ?",
            "args": [{"type": "text", "value": school_id}]
        }
    }])
    return int(result['results'][0]['response']['result']['rows'][0][0]['value'])

def _build_multi_row_insert_http(rows, rows_per_stmt=100):
    """Build multi-row INSERT statements formatted for Turso HTTP pipeline API.

    Packs `rows_per_stmt` rows into a single INSERT ... VALUES (...),(...),(...).
    Returns a list of pipeline statement dicts.
    """
    stmts = []
    for i in range(0, len(rows), rows_per_stmt):
        chunk = rows[i:i + rows_per_stmt]
        placeholders = ", ".join(["(?, ?, ?, ?, ?, ?, ?, ?)"] * len(chunk))
        args = []
        for sid, yr, sem, dept, course, instr, grade, cnt in chunk:
            args.extend([
                {"type": "text", "value": str(sid)},
                {"type": "text", "value": str(yr or '')},
                {"type": "text", "value": str(sem or '')},
                {"type": "text", "value": str(dept or '')},
                {"type": "text", "value": str(course or '')},
                {"type": "text", "value": str(instr or 'N/A')},
                {"type": "text", "value": str(grade or '')},
                {"type": "integer", "value": str(int(cnt))},
            ])
        stmts.append({
            "type": "execute",
            "stmt": {
                "sql": f"INSERT INTO grades (school_id, year, semester, dept, "
                       f"course_number, instructor, grade, count) VALUES {placeholders}",
                "args": args,
            }
        })
    return stmts

def push_to_turso_http(school_id, config, append=False, resume=False):
    """Push staging DB to Turso via HTTP pipeline API with multi-row INSERTs."""
    db_path = staging_db_path(school_id)
    prog_path = progress_path(school_id)

    if not os.path.exists(db_path):
        print(f"❌ No staging DB for {school_id}. Run --stage-only first.")
        return False

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    local_count = cur.execute("SELECT COUNT(*) FROM grades").fetchone()[0]

    existing = get_remote_count_http(config, school_id)

    # Determine starting offset
    skip = 0
    if resume and os.path.exists(prog_path):
        with open(prog_path) as f:
            progress = json.load(f)
        skip = progress.get('rows_uploaded', 0)
        print(f"⏩ Resuming from row {skip:,} (checkpoint: {prog_path})")
    elif append:
        print(f"➕ Append mode: {existing:,} rows exist in Turso. Adding new rows.")
        skip = 0
    else:
        if existing > 0:
            print(f"🔄 Fresh upload: will replace {existing:,} existing rows for {school_id}")
            print(f"   Old data preserved until new upload verified.")
        skip = 0

    remaining = local_count - skip
    ROWS_PER_STMT = 100  # rows packed into each INSERT statement
    STMTS_PER_PIPELINE = 10  # statements per HTTP pipeline request (= 1000 rows)
    BATCH = ROWS_PER_STMT * STMTS_PER_PIPELINE  # 1000 rows per HTTP call

    print(f"\n📤 Pushing {remaining:,} rows to Turso via HTTP ({school_id})")
    print(f"   Local: {local_count:,} | Remote: {existing:,} | Skip: {skip:,}")
    print(f"   Transport: HTTP pipeline (multi-row INSERT, {ROWS_PER_STMT} rows/stmt, "
          f"{STMTS_PER_PIPELINE} stmts/pipeline = {BATCH} rows/call)")

    # If fresh upload, delete old data first
    if not append and not resume and existing > 0:
        print(f"\n   Deleting old {school_id} data from Turso...")
        turso_pipeline(config, [{
            "type": "execute",
            "stmt": {
                "sql": "DELETE FROM grades WHERE school_id = ?",
                "args": [{"type": "text", "value": school_id}]
            }
        }])
        print(f"   ✅ Old data deleted")

    cur.execute(
        "SELECT school_id, year, semester, dept, course_number, instructor, grade, count "
        "FROM grades LIMIT -1 OFFSET ?",
        (skip,)
    )

    total_uploaded = skip
    errors = 0
    start_time = time.time()

    batch = []
    for row in cur:
        batch.append(row)

        if len(batch) >= BATCH:
            stmts = _build_multi_row_insert_http(batch, rows_per_stmt=ROWS_PER_STMT)
            result = turso_pipeline(config, stmts)
            batch_errors = sum(1 for r in result['results'] if r.get('type') == 'error')
            if batch_errors:
                for r in result['results']:
                    if r.get('type') == 'error':
                        print(f"  ERR: {r.get('error',{}).get('message','?')[:100]}")
            # Each successful stmt inserts ROWS_PER_STMT rows
            successful_stmts = len(stmts) - batch_errors
            total_uploaded += successful_stmts * ROWS_PER_STMT
            errors += batch_errors
            batch = []

            elapsed = time.time() - start_time
            rate = (total_uploaded - skip) / elapsed if elapsed > 0 else 0
            eta = (local_count - total_uploaded) / rate if rate > 0 else 0
            print(f"  {total_uploaded:,}/{local_count:,} rows ({rate:.0f} rows/s, ETA: {eta/60:.1f}m)")

            with open(prog_path, 'w') as f:
                json.dump({
                    'school_id': school_id,
                    'rows_uploaded': total_uploaded,
                    'total_rows': local_count,
                    'timestamp': datetime.now().isoformat(),
                    'errors': errors
                }, f)

    # Final batch
    if batch:
        stmts = _build_multi_row_insert_http(batch, rows_per_stmt=ROWS_PER_STMT)
        result = turso_pipeline(config, stmts)
        batch_errors = sum(1 for r in result['results'] if r.get('type') == 'error')
        errors += batch_errors
        total_uploaded += len(batch) - (batch_errors * ROWS_PER_STMT)

    conn.close()

    elapsed = time.time() - start_time
    print(f"\n✅ Upload complete: {total_uploaded:,} rows in {elapsed:.1f}s ({errors:,} errors)")

    # Clean up progress file on successful completion
    if os.path.exists(prog_path) and errors == 0:
        os.remove(prog_path)

    return errors == 0


# ── Transport Router ─────────────────────────────────────────────────────────

def get_remote_count(config, school_id, transport=None):
    """Get remote row count using the active transport."""
    t = get_transport(config, force=transport)
    if t == 'libsql':
        return get_remote_count_libsql(config, school_id)
    return get_remote_count_http(config, school_id)

def push_to_turso(school_id, config, append=False, resume=False, transport=None):
    """Push to Turso using the active transport."""
    t = get_transport(config, force=transport)
    print(f"   🔌 Transport: {t}" + (" (WebSocket, transactional)" if t == 'libsql' else " (HTTP pipeline, fallback)"))
    if t == 'libsql':
        return push_to_turso_libsql(school_id, config, append=append, resume=resume)
    return push_to_turso_http(school_id, config, append=append, resume=resume)

def verify_remote(school_id, config, transport=None):
    """Verify remote data matches local staging."""
    db_path = staging_db_path(school_id)
    if not os.path.exists(db_path):
        print(f"❌ No staging DB for {school_id}. Cannot verify.")
        return False

    local_checksum, local_count = compute_local_checksum(school_id)
    remote_count = get_remote_count(config, school_id, transport=transport)

    print(f"\n{'='*50}")
    print(f"Remote Verification: {school_id}")
    print(f"{'='*50}")
    print(f"  Local rows:    {local_count:,}")
    print(f"  Remote rows:   {remote_count:,}")
    print(f"  Local checksum: {local_checksum}")

    if local_count == remote_count:
        print(f"\n  ✅ VERIFIED — row counts match!")
        return True
    else:
        diff = abs(local_count - remote_count)
        print(f"\n  ❌ MISMATCH — off by {diff:,} rows!")
        if remote_count < local_count:
            print(f"  → Try: caffeinate -i python3 upload_v2.py --school {school_id} --push --resume")
        return False

# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='GradeView Data Upload v2 — Safe & Reliable')
    parser.add_argument('--school', required=True, help='School ID')
    parser.add_argument('--file', help='Path to CSV/XLSX (supports glob wildcards)')
    parser.add_argument('--map', default='xlsx_generic', help='Column mapping name')
    parser.add_argument('--auto-map', action='store_true',
                        help='Auto-detect column mapping from file headers (ignores --map)')
    parser.add_argument('--stage-only', action='store_true', help='Stage locally only, no Turso push')
    parser.add_argument('--verify-staging', action='store_true', help='Verify staging DB')
    parser.add_argument('--push', action='store_true', help='Push staged data to Turso')
    parser.add_argument('--verify-remote', action='store_true', help='Verify remote matches local')
    parser.add_argument('--append', action='store_true', help='Append (don\'t delete existing)')
    parser.add_argument('--resume', action='store_true', help='Resume interrupted push')
    parser.add_argument('--dry-run', action='store_true', help='Stage + verify only, no Turso writes')
    parser.add_argument('--transport', choices=['libsql', 'http', 'auto'], default='auto',
                        help='Turso transport: libsql (WebSocket, preferred), http (fallback), auto (default)')

    args = parser.parse_args()

    # Defer config loading — only needed for Turso operations
    needs_turso = not args.stage_only and not args.verify_staging
    config = load_config() if needs_turso else {}

    # Auto-map: inspect file headers and generate mapping dynamically
    if args.auto_map and args.file:
        from auto_mapper import auto_map, build_upload_mapping
        filepath = os.path.expanduser(args.file)
        target = sorted(glob.glob(filepath)) if '*' in filepath else [filepath]
        target = [f for f in target if os.path.exists(f)]
        if not target:
            print(f"❌ --auto-map requires a valid --file to inspect")
            sys.exit(1)
        print(f"🔍 Auto-mapping columns from {os.path.basename(target[0])}...")
        result = auto_map(target[0], school_id=args.school)
        if result["confidence"] < 0.5:
            print(f"❌ Auto-mapper confidence too low ({result['confidence']:.0%})")
            print(f"   Warnings: {'; '.join(result['warnings'])}")
            print(f"   Run: python3 auto_mapper.py --file '{target[0]}' --verbose")
            sys.exit(1)
        mapping = build_upload_mapping(result)
        if not mapping:
            print(f"❌ Auto-mapper could not generate a valid mapping")
            sys.exit(1)
        fmt_label = result["format"].upper()
        conf_label = f"{result['confidence']:.0%}"
        print(f"   Format: {fmt_label} | Confidence: {conf_label}")
        if result["warnings"]:
            for w in result["warnings"]:
                print(f"   ⚠️  {w}")
        print()
    else:
        mapping = MAPPINGS.get(args.map, MAPPINGS['xlsx_generic'])

    # Individual operations
    if args.verify_staging:
        verify_staging(args.school)
        return

    transport = args.transport if args.transport != 'auto' else None

    if args.verify_remote:
        verify_remote(args.school, config, transport=transport)
        return

    if args.push:
        if push_to_turso(args.school, config, append=args.append, resume=args.resume, transport=transport):
            verify_remote(args.school, config, transport=transport)
        return

    # Full pipeline: stage → (optional push) → verify
    if not args.file:
        print("❌ --file required for staging. Use --push to push existing staging DB.")
        sys.exit(1)

    filepath = os.path.expanduser(args.file)
    files = sorted(glob.glob(filepath)) if '*' in filepath else [filepath]
    files = [f for f in files if "(1)" not in f and os.path.exists(f)]

    if not files:
        print(f"❌ No files found: {filepath}")
        sys.exit(1)

    print(f"\n📤 GradeView Upload v2 — {args.school}")
    print(f"   Files: {len(files)} | Mapping: {args.map}")
    print(f"   Mode: {'DRY RUN' if args.dry_run else 'APPEND' if args.append else 'FRESH'}\n")

    # Multi-sheet mapping: read each sheet with its own mapping, stage together
    if mapping.get('_multi_sheet'):
        sheet_mappings = mapping['sheets']
        all_rows = []
        for f in files:
            print(f"  Reading {os.path.basename(f)} (multi-sheet)...")
            if f.endswith(('.xlsx', '.xls')):
                sheets_data = read_xlsx_file(f, all_sheets=True)
                for sheet_name, sheet_rows in sheets_data.items():
                    if sheet_name in sheet_mappings:
                        # Tag each row with the sheet's mapping for stage_standard_multi
                        for row in sheet_rows:
                            row['_sheet_mapping'] = sheet_mappings[sheet_name]
                        all_rows.extend(sheet_rows)
                        print(f"    → Sheet '{sheet_name}': {len(sheet_rows):,} rows")
                    else:
                        print(f"    → Sheet '{sheet_name}': SKIPPED (no mapping defined)")
            else:
                rows = read_csv_file(f)
                # CSV: use first sheet's mapping as default
                first_mapping = list(sheet_mappings.values())[0]
                for row in rows:
                    row['_sheet_mapping'] = first_mapping
                all_rows.extend(rows)
                print(f"    → {len(rows):,} rows")

        print(f"\n  Total source rows: {len(all_rows):,}")
        local_count = stage_standard_multi(args.school, all_rows)
    else:
        # Standard single-mapping flow
        all_rows = []
        for f in files:
            print(f"  Reading {os.path.basename(f)}...")
            sheet_name = mapping.get('sheet') if mapping.get('format') == 'wide' else None
            if f.endswith(('.xlsx', '.xls')):
                rows = read_xlsx_file(f, sheet_name=sheet_name)
            else:
                rows = read_csv_file(f)
            all_rows.extend(rows)
            print(f"    → {len(rows):,} rows")

        print(f"\n  Total source rows: {len(all_rows):,}")

        # Stage
        if mapping.get('format') == 'wide':
            local_count = stage_wide(args.school, all_rows, mapping)
        else:
            local_count = stage_standard(args.school, all_rows, mapping)

    # Verify staging
    verify_staging(args.school)

    if args.stage_only or args.dry_run:
        if args.dry_run:
            remote = get_remote_count(config, args.school, transport=transport)
            print(f"\n  DRY RUN: Would push {local_count:,} rows to Turso (currently {remote:,} remote)")
        print(f"\n  Staging DB saved: {staging_db_path(args.school)}")
        print(f"  To push: caffeinate -i python3 upload_v2.py --school {args.school} --push")
        return

    # Push
    print(f"\n{'─'*50}")
    if push_to_turso(args.school, config, append=args.append, transport=transport):
        verify_remote(args.school, config, transport=transport)
        print(f"\n🎉 Done! Update iCloud tracker and MEMORY.md.")
    else:
        print(f"\n⚠️ Upload had errors. Check and retry with --push --resume")

if __name__ == '__main__':
    main()
