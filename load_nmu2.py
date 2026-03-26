#!/usr/bin/env python3
import openpyxl
import json
import subprocess
import sys
import os

TURSO_URL = "https://gradeview-agentsolomon.aws-us-east-2.turso.io/v2/pipeline"
TURSO_TOKEN = "eyJhbGciOiJFZERTQSIsInR5cCI6IkpXVCJ9.eyJhIjoicnciLCJpYXQiOjE3NzQzMTc5NTQsImlkIjoiMDE5ZDFkOTMtZTgwMS03ODU2LThlNjYtNWY1NTgwN2I0Y2E2IiwicmlkIjoiYjBlOTI1YzAtYzM4My00ODcxLTg0NjAtYjg4OGM2NGRhNWQ3In0.uBwikoGHc1YqWDWu8HX9LIIMZ9blRUozr04x0SXBYXtyecQdcWy3RKIxXuDAAlxSHOD5R2F5k2xaQrsPk38cBA"
VALID_GRADES = {'A', 'B', 'C', 'D', 'F'}
SEMESTER_MAP = {'10': 'SPRING', '20': 'SUMMER', '30': 'FALL'}

def turso_request(requests_list):
    payload = json.dumps({"requests": requests_list})
    with open('/tmp/turso_payload.json', 'w') as f:
        f.write(payload)
    result = subprocess.run([
        'curl', '-s', '-X', 'POST', TURSO_URL,
        '-H', f'Authorization: Bearer {TURSO_TOKEN}',
        '-H', 'Content-Type: application/json',
        '-d', '@/tmp/turso_payload.json',
        '--max-time', '120'
    ], capture_output=True, text=True)
    if result.returncode != 0:
        raise Exception(f"curl error: rc={result.returncode} {result.stderr}")
    if not result.stdout.strip():
        raise Exception("Empty response from Turso")
    return json.loads(result.stdout)

def esc(s):
    return str(s).replace("'", "''") if s is not None else ''

# Step 1: Parse Excel
print("Loading Excel (may take a minute)...", flush=True)
wb = openpyxl.load_workbook('/Users/solomon/Downloads/NMU/Aikin Ryan FOIA Response Doc Grade Distribution by Term (2020-2025).xlsx')
ws = wb.active
print(f"Loaded. Rows: {ws.max_row}", flush=True)

rows = []
cur_term = None
cur_course = None
cur_instructor = None

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 2:  # skip title + header
        continue
    term_code, course, instructor, grade, count = row
    
    # Subtotal rows: course set but grade is None — skip
    if course is not None and grade is None:
        continue
    
    # Update carry-forward values
    if term_code is not None:
        cur_term = str(term_code).strip()
    if course is not None:
        cur_course = str(course).strip()
        cur_instructor = str(instructor).strip() if instructor else ''
    elif instructor is not None:
        cur_instructor = str(instructor).strip()
    
    if grade is None or count is None:
        continue
    
    grade = str(grade).strip()
    if grade not in VALID_GRADES:
        continue
    
    if cur_term is None or cur_course is None:
        continue
    
    # Parse term
    year = cur_term[:4]
    sem_code = cur_term[4:]
    semester = SEMESTER_MAP.get(sem_code)
    if semester is None:
        continue
    
    # Parse course
    parts = cur_course.split(' ', 1)
    if len(parts) != 2:
        continue
    dept, course_number = parts[0], parts[1]
    
    rows.append((year, semester, dept, course_number, cur_instructor, grade, int(count)))

print(f"Total valid rows: {len(rows)}", flush=True)

# Step 2: Batch insert
BATCH = 500
total_inserted = 0
RESUME_FROM = 3000  # already inserted
for batch_start in range(RESUME_FROM, len(rows), BATCH):
    batch = rows[batch_start:batch_start+BATCH]
    requests_list = []
    for (year, semester, dept, course_number, instructor, grade, count) in batch:
        sql = (f"INSERT INTO grades (school_id,year,semester,dept,course_number,instructor,grade,count) "
               f"VALUES ('nmu','{esc(year)}','{esc(semester)}','{esc(dept)}','{esc(course_number)}',"
               f"'{esc(instructor)}','{esc(grade)}',{count})")
        requests_list.append({"type": "execute", "stmt": {"sql": sql}})
    requests_list.append({"type": "close"})
    resp = turso_request(requests_list)
    errors = [r for r in resp.get('results', []) if r.get('type') == 'error']
    if errors:
        print(f"ERRORS in batch {batch_start}: {errors[:3]}", flush=True)
        sys.exit(1)
    total_inserted += len(batch)
    print(f"  Inserted {total_inserted}/{len(rows)}...", flush=True)

print(f"Done. Inserted {total_inserted} rows.", flush=True)

# Step 3: Count
resp = turso_request([
    {"type": "execute", "stmt": {"sql": "SELECT COUNT(*) FROM grades WHERE school_id='nmu'"}},
    {"type": "execute", "stmt": {"sql": "SELECT COUNT(*) FROM grades"}},
    {"type": "close"}
])
results = resp.get('results', [])
nmu_count = results[0]['response']['result']['rows'][0][0]['value']
total_count = results[1]['response']['result']['rows'][0][0]['value']
print(f"NMU count: {nmu_count}, Total count: {total_count}", flush=True)

# Write result
os.makedirs('/Users/solomon/.openclaw/workspace/projects/gradeview', exist_ok=True)
with open('/Users/solomon/.openclaw/workspace/projects/gradeview/nmu_load_result.md', 'w') as f:
    f.write(f"""# NMU Grade Data Load Result

**Date:** 2026-03-25  
**School:** Northern Michigan University (nmu)  
**Rows inserted:** {total_inserted}  
**NMU count in DB:** {nmu_count}  
**Total grades in DB:** {total_count}  
""")
print("Result written.", flush=True)
